"""
Admin Routes - Analytics, Reports, System Overview
"""

from flask import Blueprint, render_template, request, redirect, url_for, session, flash, send_file
from models.models import Event, Registration, Attendance, Feedback, User, Department, Venue, Role, AppConfig
from models import db
from datetime import datetime, date, timedelta
from functools import wraps
from sqlalchemy import func, or_, text
from io import BytesIO
import io
import csv
from reportlab.lib.pagesizes import landscape, letter
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.lib.units import inch
from openpyxl import Workbook, load_workbook

bp = Blueprint('admin', __name__, url_prefix='/admin')

def admin_required(f):
    """Decorator to require admin login"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session or session.get('role_name', '').lower() != 'admin':
            flash('Access denied', 'error')
            return redirect(url_for('auth.login'))
        return f(*args, **kwargs)
    return decorated_function


@bp.route('/dashboard')
@admin_required
def dashboard():
    """Admin dashboard with overview statistics"""
    # Overall statistics
    total_events = Event.query.count()
    approved_events = Event.query.filter_by(status='approved').count()
    pending_events = Event.query.filter_by(status='pending').count()
    total_students = User.query.join(User.role).filter(User.role.has(role_name='Student')).count()
    total_registrations = Registration.query.count()
    total_attendance = Attendance.query.count()
    
    # Recent events
    recent_events = Event.query.order_by(Event.created_at.desc()).limit(10).all()
    
    # Department-wise statistics
    dept_stats = db.session.query(
        Department.dept_name,
        func.count(Event.event_id).label('event_count')
    ).join(Event).group_by(Department.dept_id).all()
    
    # Upcoming events
    upcoming_events = Event.query.filter(
        Event.status == 'approved',
        Event.date >= date.today()
    ).order_by(Event.date).limit(5).all()
    
    guest_enabled = AppConfig.query.get('guest_enabled')
    return render_template('admin/dashboard.html',
                         total_events=total_events,
                         approved_events=approved_events,
                         pending_events=pending_events,
                         total_students=total_students,
                         total_registrations=total_registrations,
                         total_attendance=total_attendance,
                         recent_events=recent_events,
                         dept_stats=dept_stats,
                         upcoming_events=upcoming_events,
                         guest_enabled=guest_enabled)


@bp.route('/events')
@admin_required
def events():
    """View all events with filters"""
    # Get filter parameters
    status_filter = request.args.get('status', '')
    dept_filter = request.args.get('department', '')
    organizer_filter = request.args.get('organizer', '')
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    
    # Build query
    query = Event.query
    
    if status_filter:
        query = query.filter_by(status=status_filter)
    
    if dept_filter:
        query = query.filter_by(dept_id=int(dept_filter))

    if organizer_filter:
        query = query.filter_by(organizer_id=int(organizer_filter))
    
    if date_from:
        query = query.filter(Event.date >= datetime.strptime(date_from, '%Y-%m-%d').date())
    
    if date_to:
        query = query.filter(Event.date <= datetime.strptime(date_to, '%Y-%m-%d').date())
    
    events = query.order_by(Event.date.desc()).all()
    
    # Get departments for filter
    departments = Department.query.all()
    organizers = User.query.join(User.role).filter(
        or_(
            User.role.has(role_name='Event Organizer'),
            User.role.has(role_name='Organizer')
        )
    ).order_by(User.full_name.asc()).all()
    
    return render_template('admin/events.html',
                         events=events,
                         departments=departments,
                         organizers=organizers,
                         status_filter=status_filter,
                         dept_filter=dept_filter,
                         organizer_filter=organizer_filter,
                         date_from=date_from,
                         date_to=date_to)


@bp.route('/events/export')
@admin_required
def export_events():
    """Export filtered events to Excel or PDF"""
    status_filter = request.args.get('status', '')
    dept_filter = request.args.get('department', '')
    organizer_filter = request.args.get('organizer', '')
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    export_format = (request.args.get('format') or 'xlsx').lower()

    query = Event.query

    if status_filter:
        query = query.filter_by(status=status_filter)

    if dept_filter:
        query = query.filter_by(dept_id=int(dept_filter))

    if organizer_filter:
        query = query.filter_by(organizer_id=int(organizer_filter))

    if date_from:
        query = query.filter(Event.date >= datetime.strptime(date_from, '%Y-%m-%d').date())

    if date_to:
        query = query.filter(Event.date <= datetime.strptime(date_to, '%Y-%m-%d').date())

    events = query.order_by(Event.date.desc()).all()

    if export_format == 'pdf':
        buffer = BytesIO()
        c = canvas.Canvas(buffer, pagesize=landscape(letter))
        width, height = landscape(letter)

        c.setFont('Helvetica-Bold', 16)
        c.drawString(0.6 * inch, height - 0.6 * inch, 'Events Report')

        headers = ['Title', 'Date', 'Status', 'Department', 'Organizer']
        col_widths = [3.2 * inch, 1.4 * inch, 1.2 * inch, 2.0 * inch, 2.0 * inch]
        start_x = 0.6 * inch
        start_y = height - 1.1 * inch
        row_height = 0.35 * inch

        c.setFillColor(colors.HexColor('#1f2937'))
        c.setFont('Helvetica-Bold', 10)
        x = start_x
        for idx, header in enumerate(headers):
            c.drawString(x + 4, start_y, header)
            x += col_widths[idx]

        c.setStrokeColor(colors.HexColor('#e5e7eb'))
        y = start_y - row_height
        c.setFont('Helvetica', 9)
        for event in events:
            if y < 0.7 * inch:
                c.showPage()
                y = height - 0.8 * inch
            values = [
                event.title,
                event.date.strftime('%Y-%m-%d'),
                event.status,
                event.department.dept_name if event.department else '—',
                event.organizer.full_name if event.organizer else '—'
            ]
            x = start_x
            for idx, value in enumerate(values):
                c.drawString(x + 4, y, str(value))
                x += col_widths[idx]
            y -= row_height

        c.save()
        buffer.seek(0)
        return send_file(
            buffer,
            as_attachment=True,
            download_name='events_report.pdf',
            mimetype='application/pdf'
        )

    wb = Workbook()
    ws = wb.active
    ws.title = 'Events'
    ws.append(['Title', 'Date', 'Status', 'Department', 'Organizer'])
    for event in events:
        ws.append([
            event.title,
            event.date.strftime('%Y-%m-%d'),
            event.status,
            event.department.dept_name if event.department else '—',
            event.organizer.full_name if event.organizer else '—'
        ])

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        as_attachment=True,
        download_name='events_report.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@bp.route('/event/<int:event_id>')
@admin_required
def view_event(event_id):
    """View detailed event information"""
    event = Event.query.get_or_404(event_id)
    
    # Get registrations
    registrations = Registration.query.filter_by(event_id=event_id).all()
    
    # Get attendance
    attendance_records = db.session.query(Attendance).join(Registration).filter(
        Registration.event_id == event_id
    ).all()
    
    # Get feedback
    feedback_records = Feedback.query.filter_by(event_id=event_id).all()
    
    # Calculate average rating
    avg_rating = db.session.query(func.avg(Feedback.rating)).filter_by(
        event_id=event_id
    ).scalar() or 0
    
    return render_template('admin/view_event.html',
                         event=event,
                         registrations=registrations,
                         attendance_records=attendance_records,
                         feedback_records=feedback_records,
                         avg_rating=round(avg_rating, 2))


@bp.route('/reports')
@admin_required
def reports():
    """Generate various reports"""
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    dept_filter = request.args.get('dept_id', '')

    event_filters = []
    if date_from:
        event_filters.append(Event.date >= datetime.strptime(date_from, '%Y-%m-%d').date())
    if date_to:
        event_filters.append(Event.date <= datetime.strptime(date_to, '%Y-%m-%d').date())
    if dept_filter:
        try:
            event_filters.append(Event.dept_id == int(dept_filter))
        except ValueError:
            dept_filter = ''

    # Monthly event statistics (filtered)
    current_month = date.today().replace(day=1)
    events_this_month = Event.query.filter(
        Event.date >= current_month,
        *event_filters
    ).count()
    
    # Student participation statistics
    top_students = db.session.query(
        User.full_name,
        User.email,
        func.count(Registration.registration_id).label('event_count')
    ).join(Registration).join(Event, Registration.event_id == Event.event_id).filter(
        User.role.has(role_name='Student'),
        *event_filters
    ).group_by(User.user_id).order_by(func.count(Registration.registration_id).desc()).limit(10).all()
    
    # Event organizer statistics
    top_organizers = db.session.query(
        User.full_name,
        User.email,
        func.count(Event.event_id).label('event_count')
    ).join(Event, User.user_id == Event.organizer_id).filter(
        *event_filters
    ).group_by(User.user_id).order_by(func.count(Event.event_id).desc()).limit(10).all()

    organizer_feedback = db.session.query(
        User.full_name,
        User.email,
        func.avg(Feedback.rating).label('avg_rating'),
        func.count(Feedback.feedback_id).label('feedback_count')
    ).join(Event, Event.organizer_id == User.user_id).join(
        Feedback, Feedback.event_id == Event.event_id
    ).filter(
        *event_filters
    ).group_by(User.user_id).order_by(func.avg(Feedback.rating).desc()).limit(10).all()
    
    # Department-wise participation
    dept_participation = db.session.query(
        Department.dept_name,
        func.count(Registration.registration_id).label('registration_count')
    ).join(User, User.dept_id == Department.dept_id).join(
        Registration, Registration.student_id == User.user_id
    ).join(Event, Registration.event_id == Event.event_id).filter(
        *event_filters
    ).group_by(Department.dept_id).all()
    
    # Feedback summary
    feedback_summary = db.session.query(
        Feedback.rating,
        func.count(Feedback.feedback_id).label('count')
    ).join(Event, Feedback.event_id == Event.event_id).filter(
        *event_filters
    ).group_by(Feedback.rating).all()
    
    return render_template('admin/reports.html',
                         events_this_month=events_this_month,
                         top_students=top_students,
                         top_organizers=top_organizers,
                         organizer_feedback=organizer_feedback,
                         dept_participation=dept_participation,
                         feedback_summary=feedback_summary,
                         departments=Department.query.order_by(Department.dept_name.asc()).all(),
                         date_from=date_from,
                         date_to=date_to,
                         dept_filter=dept_filter)


@bp.route('/reports/export')
@admin_required
def export_reports():
    """Export reports to Excel or PDF"""
    export_format = (request.args.get('format') or 'xlsx').lower()
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    dept_filter = request.args.get('dept_id', '')

    event_filters = []
    if date_from:
        event_filters.append(Event.date >= datetime.strptime(date_from, '%Y-%m-%d').date())
    if date_to:
        event_filters.append(Event.date <= datetime.strptime(date_to, '%Y-%m-%d').date())
    if dept_filter:
        try:
            event_filters.append(Event.dept_id == int(dept_filter))
        except ValueError:
            dept_filter = ''

    top_students = db.session.query(
        User.full_name,
        User.email,
        func.count(Registration.registration_id).label('event_count')
    ).join(Registration).join(Event, Registration.event_id == Event.event_id).filter(
        User.role.has(role_name='Student'),
        *event_filters
    ).group_by(User.user_id).order_by(func.count(Registration.registration_id).desc()).limit(10).all()

    top_organizers = db.session.query(
        User.full_name,
        User.email,
        func.count(Event.event_id).label('event_count')
    ).join(Event, User.user_id == Event.organizer_id).filter(
        *event_filters
    ).group_by(User.user_id).order_by(func.count(Event.event_id).desc()).limit(10).all()

    organizer_feedback = db.session.query(
        User.full_name,
        User.email,
        func.avg(Feedback.rating).label('avg_rating'),
        func.count(Feedback.feedback_id).label('feedback_count')
    ).join(Event, Event.organizer_id == User.user_id).join(
        Feedback, Feedback.event_id == Event.event_id
    ).filter(
        *event_filters
    ).group_by(User.user_id).order_by(func.avg(Feedback.rating).desc()).limit(10).all()

    if export_format == 'pdf':
        buffer = BytesIO()
        c = canvas.Canvas(buffer, pagesize=landscape(letter))
        width, height = landscape(letter)
        y = height - 0.6 * inch

        c.setFont('Helvetica-Bold', 16)
        c.drawString(0.6 * inch, y, 'Reports & Analytics')
        y -= 0.5 * inch

        def draw_table(title, headers, rows):
            nonlocal y
            if y < 1.4 * inch:
                c.showPage()
                y = height - 0.8 * inch
            c.setFont('Helvetica-Bold', 12)
            c.drawString(0.6 * inch, y, title)
            y -= 0.3 * inch

            c.setFont('Helvetica-Bold', 9)
            x = 0.6 * inch
            col_widths = [2.6 * inch, 2.6 * inch, 1.6 * inch, 1.6 * inch]
            for idx, header in enumerate(headers):
                c.drawString(x + 2, y, header)
                x += col_widths[idx]
            y -= 0.3 * inch

            c.setFont('Helvetica', 9)
            for row in rows:
                if y < 0.8 * inch:
                    c.showPage()
                    y = height - 0.8 * inch
                x = 0.6 * inch
                for idx, value in enumerate(row):
                    c.drawString(x + 2, y, str(value))
                    x += col_widths[idx]
                y -= 0.28 * inch

            y -= 0.2 * inch

        draw_table(
            'Top Participating Students',
            ['Name', 'Email', 'Events Attended', ''],
            [(s[0], s[1], s[2], '') for s in top_students]
        )

        draw_table(
            'Most Active Organizers',
            ['Name', 'Email', 'Events Created', ''],
            [(o[0], o[1], o[2], '') for o in top_organizers]
        )

        draw_table(
            'Best Rated Organizers (Feedback)',
            ['Name', 'Email', 'Avg Rating', 'Feedback Count'],
            [(o[0], o[1], f"{(o[2] or 0):.2f}", o[3]) for o in organizer_feedback]
        )

        c.save()
        buffer.seek(0)
        return send_file(
            buffer,
            as_attachment=True,
            download_name='reports_analytics.pdf',
            mimetype='application/pdf'
        )

    wb = Workbook()
    ws_students = wb.active
    ws_students.title = 'Top Students'
    ws_students.append(['Name', 'Email', 'Events Attended'])
    for student in top_students:
        ws_students.append([student[0], student[1], student[2]])

    ws_org = wb.create_sheet('Active Organizers')
    ws_org.append(['Name', 'Email', 'Events Created'])
    for org in top_organizers:
        ws_org.append([org[0], org[1], org[2]])

    ws_feedback = wb.create_sheet('Best Organizers')
    ws_feedback.append(['Name', 'Email', 'Avg Rating', 'Feedback Count'])
    for org in organizer_feedback:
        ws_feedback.append([org[0], org[1], float(org[2] or 0), org[3]])

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        as_attachment=True,
        download_name='reports_analytics.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@bp.route('/feedback')
@admin_required
def feedback():
    """View feedback summary by event with filters"""
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    dept_filter = request.args.get('dept_id', '')
    event_filter = request.args.get('event_id', '')
    organizer_filter = request.args.get('organizer_id', '')

    event_filters = []
    if date_from:
        event_filters.append(Event.date >= datetime.strptime(date_from, '%Y-%m-%d').date())
    if date_to:
        event_filters.append(Event.date <= datetime.strptime(date_to, '%Y-%m-%d').date())
    if dept_filter:
        try:
            event_filters.append(Event.dept_id == int(dept_filter))
        except ValueError:
            dept_filter = ''
    if event_filter:
        try:
            event_filters.append(Event.event_id == int(event_filter))
        except ValueError:
            event_filter = ''
    if organizer_filter:
        try:
            event_filters.append(Event.organizer_id == int(organizer_filter))
        except ValueError:
            organizer_filter = ''

    summary = db.session.query(
        Event,
        func.avg(Feedback.rating).label('avg_rating'),
        func.count(Feedback.feedback_id).label('feedback_count')
    ).join(Feedback, Feedback.event_id == Event.event_id).filter(
        *event_filters
    ).group_by(Event.event_id).order_by(Event.date.desc()).all()

    total_feedback = Feedback.query.count()
    avg_rating = db.session.query(func.avg(Feedback.rating)).scalar() or 0

    departments = Department.query.order_by(Department.dept_name.asc()).all()
    events = Event.query.order_by(Event.date.desc()).all()
    organizers = User.query.join(User.role).filter(
        or_(
            User.role.has(role_name='Event Organizer'),
            User.role.has(role_name='Organizer')
        )
    ).order_by(User.full_name.asc()).all()

    return render_template(
        'admin/feedback.html',
        summary=summary,
        total_feedback=total_feedback,
        avg_rating=round(avg_rating, 2),
        departments=departments,
        events=events,
        date_from=date_from,
        date_to=date_to,
        dept_filter=dept_filter,
        event_filter=event_filter,
        organizer_filter=organizer_filter,
        organizers=organizers
    )


@bp.route('/feedback/event/<int:event_id>')
@admin_required
def feedback_event(event_id):
    """View detailed feedback for a specific event"""
    event = Event.query.get_or_404(event_id)
    feedback_items = Feedback.query.filter_by(event_id=event_id).order_by(Feedback.submitted_at.desc()).all()
    avg_rating = db.session.query(func.avg(Feedback.rating)).filter_by(event_id=event_id).scalar() or 0
    return render_template('admin/feedback_detail.html',
                         event=event,
                         feedback_items=feedback_items,
                         avg_rating=round(avg_rating, 2))


@bp.route('/users')
@admin_required
def users():
    """View all users"""
    search = (request.args.get('q') or '').strip()
    role_filter = request.args.get('role_id')
    dept_filter = request.args.get('dept_id')

    query = User.query

    if search:
        like = f"%{search}%"
        query = query.filter(or_(User.full_name.ilike(like), User.email.ilike(like), User.username.ilike(like)))

    if role_filter:
        try:
            query = query.filter(User.role_id == int(role_filter))
        except ValueError:
            pass

    if dept_filter:
        try:
            query = query.filter(User.dept_id == int(dept_filter))
        except ValueError:
            pass

    users = query.order_by(User.created_at.desc()).all()
    departments = Department.query.order_by(Department.dept_name.asc()).all()
    roles = Role.query.order_by(Role.role_name.asc()).all()
    return render_template(
        'admin/users.html',
        users=users,
        departments=departments,
        roles=roles,
        search=search,
        role_filter=role_filter,
        dept_filter=dept_filter
    )


@bp.route('/users/create', methods=['POST'])
@admin_required
def create_user():
    """Create a single organizer or HOD user"""
    full_name = (request.form.get('full_name') or '').strip()
    username = (request.form.get('username') or '').strip()
    email = (request.form.get('email') or '').strip().lower()
    password = request.form.get('password') or ''
    role_id = request.form.get('role_id')
    dept_id = request.form.get('dept_id')

    if not full_name or not email or not password or not role_id:
        flash('Please fill all required fields for user creation.', 'error')
        return redirect(url_for('admin.users'))

    # Enforce minimum password length
    if len(password) < 8:
        flash('Password must be at least 8 characters long.', 'error')
        return redirect(url_for('admin.users'))

    role = Role.query.get(int(role_id)) if role_id else None
    if not role:
        flash('Invalid role selected.', 'error')
        return redirect(url_for('admin.users'))
    # Admin accounts cannot be created via this interface
    if (role.role_name or '').strip().lower() == 'admin':
        flash('Cannot create Admin accounts via this interface.', 'error')
        return redirect(url_for('admin.users'))

    if User.query.filter_by(email=email).first():
        flash('Email already exists. Please use a different email.', 'error')
        return redirect(url_for('admin.users'))

    if username and User.query.filter_by(username=username).first():
        flash('Username already exists. Please use a different username.', 'error')
        return redirect(url_for('admin.users'))

    # If creating HOD, dept is required and only one HOD per department allowed
    if (role.role_name or '').strip().lower() == 'hod':
        if not dept_id:
            flash('Please select a department for HOD.', 'error')
            return redirect(url_for('admin.users'))
        try:
            dept_int = int(dept_id)
        except Exception:
            flash('Invalid department selected for HOD.', 'error')
            return redirect(url_for('admin.users'))
        existing_hod = User.query.join(Role).filter(Role.role_name.ilike('hod'), User.dept_id == dept_int).first()
        if existing_hod:
            flash('A HOD already exists for the selected department.', 'error')
            return redirect(url_for('admin.users'))
    # If creating Principal, ensure only one Principal exists
    if (role.role_name or '').strip().lower() == 'principal':
        existing_principal = User.query.join(Role).filter(Role.role_name.ilike('principal')).first()
        if existing_principal:
            flash('A Principal account already exists.', 'error')
            return redirect(url_for('admin.users'))

    user = User(
        full_name=full_name,
        username=username or None,
        email=email,
        role_id=role.role_id,
        dept_id=int(dept_id) if dept_id else None
    )
    user.set_password(password)
    db.session.add(user)
    db.session.commit()

    flash('User created successfully.', 'success')
    return redirect(url_for('admin.users'))


@bp.route('/users/<int:user_id>/edit', methods=['GET', 'POST'])
@admin_required
def edit_user(user_id):
    """Edit an existing user's profile details."""
    user = User.query.get_or_404(user_id)
    departments = Department.query.order_by(Department.dept_name.asc()).all()
    roles = Role.query.order_by(Role.role_name.asc()).all()

    if request.method == 'POST':
        full_name = (request.form.get('full_name') or '').strip()
        username = (request.form.get('username') or '').strip()
        email = (request.form.get('email') or '').strip().lower()
        role_id = request.form.get('role_id')
        dept_id = request.form.get('dept_id')

        if not full_name or not email or not role_id:
            flash('Full name, email, and role are required.', 'error')
            return redirect(url_for('admin.edit_user', user_id=user_id))

        role = Role.query.get(int(role_id)) if role_id else None
        if not role:
            flash('Invalid role selected.', 'error')
            return redirect(url_for('admin.edit_user', user_id=user_id))

        # Prevent assigning Admin role here
        if (role.role_name or '').strip().lower() == 'admin':
            flash('Cannot assign Admin role via this interface.', 'error')
            return redirect(url_for('admin.edit_user', user_id=user_id))

        # If assigning HOD, dept required and ensure no other HOD for that dept
        if (role.role_name or '').strip().lower() == 'hod':
            if not dept_id:
                flash('Please select a department for HOD.', 'error')
                return redirect(url_for('admin.edit_user', user_id=user_id))
            try:
                dept_int = int(dept_id)
            except Exception:
                flash('Invalid department selected for HOD.', 'error')
                return redirect(url_for('admin.edit_user', user_id=user_id))
            existing_hod = User.query.join(Role).filter(Role.role_name.ilike('hod'), User.dept_id == dept_int, User.user_id != user.user_id).first()
            if existing_hod:
                flash('A HOD already exists for the selected department.', 'error')
                return redirect(url_for('admin.edit_user', user_id=user_id))

        # If assigning Principal, ensure only one Principal exists (excluding this user)
        if (role.role_name or '').strip().lower() == 'principal':
            existing_principal = User.query.join(Role).filter(Role.role_name.ilike('principal'), User.user_id != user.user_id).first()
            if existing_principal:
                flash('A Principal account already exists.', 'error')
                return redirect(url_for('admin.edit_user', user_id=user_id))

        if username and username != user.username:
            if User.query.filter(User.username == username, User.user_id != user.user_id).first():
                flash('Username already exists. Please use a different username.', 'error')
                return redirect(url_for('admin.edit_user', user_id=user_id))

        if email != user.email:
            if User.query.filter(User.email == email, User.user_id != user.user_id).first():
                flash('Email already exists. Please use a different email.', 'error')
                return redirect(url_for('admin.edit_user', user_id=user_id))

        user.full_name = full_name
        user.username = username or None
        user.email = email
        user.role_id = role.role_id
        user.dept_id = int(dept_id) if dept_id else None

        db.session.commit()
        flash('User updated successfully.', 'success')
        return redirect(url_for('admin.users'))

    return render_template('admin/edit_user.html', user=user, roles=roles, departments=departments)


@bp.route('/users/<int:user_id>/delete', methods=['POST'])
@admin_required
def delete_user(user_id):
    """Delete a user if safe to remove."""
    if session.get('user_id') == user_id:
        flash('You cannot delete your own account.', 'error')
        return redirect(url_for('admin.users'))

    user = User.query.get_or_404(user_id)

    if user.role and user.role.role_name and user.role.role_name.lower() == 'admin':
        flash('Admin accounts cannot be deleted.', 'error')
        return redirect(url_for('admin.users'))

    # Prevent deleting other admin accounts
    if user.role and user.role.role_name and user.role.role_name.lower() == 'admin':
        flash('Admin accounts cannot be deleted.', 'error')
        return redirect(url_for('admin.users'))

    # Conservative cleanup of related records to allow deletion
    try:
        # Delete attendance scanned by user
        db.session.execute(text('DELETE FROM attendance WHERE scanned_by = :uid'), {'uid': user_id})

        # Delete team invitations where invitee is this user
        db.session.execute(text('DELETE FROM team_invitations WHERE invitee_id = :uid'), {'uid': user_id})

        # Delete registrations for this user
        db.session.execute(text('DELETE FROM registrations WHERE student_id = :uid'), {'uid': user_id})

        # Delete teams led by this user
        db.session.execute(text('DELETE FROM teams WHERE leader_id = :uid'), {'uid': user_id})

        # Delete certificates issued to this user
        db.session.execute(text('DELETE FROM certificates WHERE student_id = :uid'), {'uid': user_id})

        # Delete feedback by this user
        db.session.execute(text('DELETE FROM feedback WHERE student_id = :uid'), {'uid': user_id})

        # Delete approvals by this user
        db.session.execute(text('DELETE FROM approvals WHERE approver_id = :uid'), {'uid': user_id})

        # Delete events organized by this user (this will cascade to registrations/teams if DB is configured)
        db.session.execute(text('DELETE FROM events WHERE organizer_id = :uid'), {'uid': user_id})

        # Finally delete the user
        db.session.delete(user)
        db.session.commit()
        flash('User deleted successfully.', 'success')
        return redirect(url_for('admin.users'))
    except Exception as e:
        db.session.rollback()
        current_app.logger.exception('Failed to delete user %s', user_id)
        flash(f'Failed to delete user: {e}', 'error')
        return redirect(url_for('admin.users'))





def _normalize_headers(headers):
    return [str(h or '').strip().lower().replace(' ', '_').replace('-', '_') for h in headers]


def _resolve_department(dept_id, dept_name, dept_map):
    if dept_id:
        try:
            return int(dept_id)
        except (TypeError, ValueError):
            return None
    if dept_name:
        return dept_map.get(str(dept_name).strip().lower())
    return None


@bp.route('/toggle-guest-login', methods=['POST'])
@admin_required
def toggle_guest_login():
    """Toggle the global guest login feature on/off via `app_config.guest_enabled`."""
    cfg = AppConfig.query.get('guest_enabled')
    if not cfg:
        cfg = AppConfig(key='guest_enabled', value='1')
        db.session.add(cfg)

    current = (cfg.value or '').strip()
    cfg.value = '0' if current == '1' else '1'
    db.session.commit()
    flash(f"Guest login {'enabled' if cfg.value == '1' else 'disabled'}.", 'success')
    # Return admin to the dashboard where the toggle was available
    return redirect(url_for('admin.dashboard'))


@bp.route('/users/bulk-upload', methods=['POST'])
@admin_required
def bulk_upload_students():
    """Bulk create student users from CSV/XLSX"""
    upload = request.files.get('file')
    default_password = request.form.get('default_password') or ''

    if not upload or not upload.filename:
        flash('Please select a CSV or XLSX file to upload.', 'error')
        return redirect(url_for('admin.users'))

    filename = upload.filename.lower()
    if not (filename.endswith('.csv') or filename.endswith('.xlsx')):
        flash('Unsupported file type. Please upload a CSV or XLSX file.', 'error')
        return redirect(url_for('admin.users'))

    student_role = Role.query.filter_by(role_name='Student').first()
    if not student_role:
        flash('Student role not found in database.', 'error')
        return redirect(url_for('admin.users'))

    departments = Department.query.all()
    dept_map = {d.dept_name.strip().lower(): d.dept_id for d in departments}
    dept_name_map = {d.dept_id: d.dept_name for d in departments}

    created = 0
    skipped = 0
    errors = []
    created_entries = []

    def process_row(row):
        nonlocal created, skipped, errors, created_entries
        full_name = (row.get('full_name') or '').strip()
        username = (row.get('username') or row.get('reg_no') or row.get('registration_no') or '').strip()
        if not username:
            username = None
        email = (row.get('email') or '').strip().lower()
        password = (row.get('password') or '').strip()
        dept_id = row.get('dept_id')
        dept_name = row.get('dept_name') or row.get('department')

        if not full_name or not email:
            skipped += 1
            errors.append('Missing full_name or email')
            return

        if User.query.filter_by(email=email).first():
            skipped += 1
            return

        if username and User.query.filter_by(username=username).first():
            skipped += 1
            errors.append(f'Duplicate username for {email}')
            return

        if not password:
            password = default_password.strip()

        if not password:
            skipped += 1
            errors.append(f'Missing password for {email}')
            return
        # Enforce minimum password length
        if len(password) < 8:
            skipped += 1
            errors.append(f'Password too short for {email} (min 8 chars)')
            return

        resolved_dept_id = _resolve_department(dept_id, dept_name, dept_map)

        user = User(
            full_name=full_name,
            username=username or None,
            email=email,
            role_id=student_role.role_id,
            dept_id=resolved_dept_id
        )
        user.set_password(password)
        db.session.add(user)
        created += 1
        created_entries.append({
            'full_name': full_name,
            'username': username or '',
            'email': email,
            'password': password,
            'dept_name': dept_name_map.get(resolved_dept_id, dept_name or '')
        })

    try:
        if filename.endswith('.csv'):
            content = upload.read().decode('utf-8', errors='ignore')
            reader = csv.DictReader(io.StringIO(content))
            for row in reader:
                process_row(row)
        else:
            wb = load_workbook(upload, read_only=True, data_only=True)
            ws = wb.active
            rows = ws.iter_rows(values_only=True)
            headers = next(rows, None)
            if not headers:
                flash('The uploaded file is empty.', 'error')
                return redirect(url_for('admin.users'))
            normalized = _normalize_headers(headers)
            for row in rows:
                row_dict = {normalized[idx]: (row[idx] if idx < len(row) else None) for idx in range(len(normalized))}
                process_row(row_dict)
    except Exception as exc:
        flash(f'Failed to process file: {exc}', 'error')
        return redirect(url_for('admin.users'))

    db.session.commit()

    if created == 0:
        message = f'Bulk upload completed. Created: {created}, Skipped: {skipped}.'
        if errors:
            message += ' Some rows had issues.'
        flash(message, 'warning')
        if errors:
            flash('Errors: ' + '; '.join(errors[:5]), 'warning')
        return redirect(url_for('admin.users'))

    wb = Workbook()
    ws = wb.active
    ws.title = 'Created Students'
    ws.append(['Full Name', 'Username', 'Email', 'Password', 'Department'])
    for entry in created_entries:
        ws.append([
            entry['full_name'],
            entry['username'],
            entry['email'],
            entry['password'],
            entry['dept_name']
        ])

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    message = f'Bulk upload completed. Created: {created}, Skipped: {skipped}.'
    if errors:
        message += ' Some rows had issues.'
    flash(message, 'success')
    if errors:
        flash('Errors: ' + '; '.join(errors[:5]), 'warning')

    return send_file(
        output,
        as_attachment=True,
        download_name='created_student_credentials.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@bp.route('/users/bulk-template')
@admin_required
def bulk_template():
    """Download CSV template for bulk student upload, prefilled with department name."""
    dept_id = request.args.get('dept_id')
    dept_name = 'Department'

    if dept_id:
        try:
            dept = Department.query.get(int(dept_id))
            if dept:
                dept_name = dept.dept_name
        except ValueError:
            pass

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['full_name', 'username', 'email', 'dept_name', 'password'])
    writer.writerow(['Student Name', '22CS001', 'student@campus.edu', dept_name, 'Student@123'])

    csv_bytes = io.BytesIO(output.getvalue().encode('utf-8'))
    csv_bytes.seek(0)

    filename = f"student_template_{dept_name.replace(' ', '_')}.csv"
    return send_file(
        csv_bytes,
        as_attachment=True,
        download_name=filename,
        mimetype='text/csv'
    )

# --- Guest management added by feature: time-limited guest accounts ---
from models.models import AppConfig

@bp.route('/guests')
@admin_required
def guests():
    """Guest management dashboard"""
    guest_enabled = AppConfig.query.get('guest_enabled')
    guest_validity = AppConfig.query.get('guest_validity_days')
    cleanup_policy = AppConfig.query.get('guest_cleanup_policy')

    # Prefer role-based guest detection; join roles to find users with Guest role
    guests = User.query.join(Role).filter(Role.role_name.ilike('guest')).order_by(User.created_at.desc()).all()
    return render_template('admin/guests.html', guests=guests, guest_enabled=guest_enabled, guest_validity=guest_validity, cleanup_policy=cleanup_policy)

@bp.route('/guests/update_settings', methods=['POST'])
@admin_required
def guests_update_settings():
    enabled = request.form.get('guest_enabled') or '0'
    validity = request.form.get('guest_validity_days') or '30'
    policy = request.form.get('guest_cleanup_policy') or 'archive'
    def setk(k,v):
        s = AppConfig.query.get(k)
        if not s:
            s = AppConfig(key=k, value=v)
            db.session.add(s)
        else:
            s.value = v
    setk('guest_enabled', '1' if enabled=='1' else '0')
    setk('guest_validity_days', str(int(validity)))
    setk('guest_cleanup_policy', policy)
    db.session.commit()
    flash('Guest settings updated', 'success')
    return redirect(url_for('admin.guests'))

@bp.route('/guests/deactivate/<int:user_id>', methods=['POST'])
@admin_required
def guests_deactivate(user_id):
    user = User.query.get_or_404(user_id)
    is_guest_user = (user.role and (user.role.role_name or '').strip().lower() == 'guest')
    if not is_guest_user:
        flash('Not a guest user', 'error')
        return redirect(url_for('admin.guests'))
    user.guest_status = 'disabled'
    db.session.commit()
    flash('Guest deactivated', 'success')
    return redirect(url_for('admin.guests'))

@bp.route('/guests/delete/<int:user_id>', methods=['POST'])
@admin_required
def guests_delete(user_id):
    user = User.query.get_or_404(user_id)
    is_guest_user = (user.role and (user.role.role_name or '').strip().lower() == 'guest')
    if not is_guest_user:
        flash('Not a guest user', 'error')
        return redirect(url_for('admin.guests'))
    # Delete related data: registrations, attendance, certificates, feedback
    Registration.query.filter_by(student_id=user.user_id).delete()
    Attendance.query.filter(Attendance.scanned_by==user.user_id).delete()
    Certificate.query.filter_by(student_id=user.user_id).delete()
    Feedback.query.filter_by(student_id=user.user_id).delete()
    db.session.delete(user)
    db.session.commit()
    flash('Guest and related data deleted', 'success')
    return redirect(url_for('admin.guests'))

@bp.route('/guests/cleanup', methods=['POST'])
@admin_required
def guests_cleanup():
    # Run cleanup: deactivate expired guests and delete/ archive based on policy
    from datetime import datetime
    now = datetime.utcnow()
    # Find expired users by Guest role (preferred) or legacy flag
    expired = User.query.join(Role).filter(Role.role_name.ilike('guest'), User.expiry_date!=None, User.expiry_date<now).all()
    policy = (AppConfig.query.get('guest_cleanup_policy').value if AppConfig.query.get('guest_cleanup_policy') else 'archive')
    for u in expired:
        u.guest_status = 'expired'
        if policy == 'delete':
            Registration.query.filter_by(student_id=u.user_id).delete()
            Attendance.query.filter(Attendance.scanned_by==u.user_id).delete()
            Certificate.query.filter_by(student_id=u.user_id).delete()
            Feedback.query.filter_by(student_id=u.user_id).delete()
            db.session.delete(u)
    db.session.commit()
    flash('Guest cleanup completed', 'success')
    return redirect(url_for('admin.guests'))
