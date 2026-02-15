"""
Event Organizer Routes - Create Events, Manage Registrations, QR Scanning
"""

from flask import Blueprint, render_template, request, redirect, url_for, session, flash, jsonify, current_app, send_file
import os
import json
from models.models import Event, Venue, Department, Registration, Attendance, User, Approval, Certificate, Feedback, CertificateTemplate
from models import db
from datetime import datetime, date, timedelta
import uuid
from functools import wraps
from sqlalchemy import func
from utils.certificate_generator import generate_certificate, generate_certificate_with_template
from utils.email_utils import send_email
from utils.qr_utils import validate_qr_code
from werkzeug.utils import secure_filename
from openpyxl import load_workbook
from io import BytesIO
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.lib.units import mm

bp = Blueprint('organizer', __name__, url_prefix='/organizer')

def organizer_required(f):
    """Decorator to require organizer login"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session or session.get('role_name', '').lower() not in ('event organizer', 'organizer'):
            flash('Access denied', 'error')
            return redirect(url_for('auth.login'))
        return f(*args, **kwargs)
    return decorated_function


@bp.route('/dashboard')
@organizer_required
def dashboard():
    """Organizer dashboard"""
    organizer_id = session['user_id']
    search = (request.args.get('q') or '').strip()
    page = request.args.get('page', 1, type=int)
    
    # Detect mobile via User-Agent for different pagination limits
    user_agent = request.headers.get('User-Agent', '').lower()
    is_mobile = any(x in user_agent for x in ['mobile', 'android', 'iphone', 'ipad', 'ipod'])
    per_page = 5 if is_mobile else 10
    
    # Get organizer's events with pagination
    events_query = Event.query.filter_by(organizer_id=organizer_id)
    if search:
        events_query = events_query.filter(Event.title.ilike(f"%{search}%"))
    pagination = events_query.order_by(Event.created_at.desc()).paginate(page=page, per_page=per_page, error_out=False)
    events = pagination.items
    
    # Count statistics
    pending_count = Event.query.filter_by(organizer_id=organizer_id, status='pending').count()
    approved_count = Event.query.filter_by(organizer_id=organizer_id, status='approved').count()
    rejected_count = Event.query.filter_by(organizer_id=organizer_id, status='rejected').count()

    past_event_ids = set()
    now = datetime.now()
    for ev in events:
        try:
            end_dt = datetime.combine(ev.date, ev.end_time)
            if now >= end_dt:
                past_event_ids.add(ev.event_id)
        except Exception:
            pass
    
    # Build notifications for organizer
    notifications = []
    
    # Get recent approvals (last 7 days) for organizer's events
    from datetime import timedelta
    week_ago = now - timedelta(days=7)
    
    recent_approvals = Approval.query.join(Event).filter(
        Event.organizer_id == organizer_id,
        Approval.approved_at != None,
        Approval.approved_at >= week_ago
    ).order_by(Approval.approved_at.desc()).limit(5).all()
    
    for approval in recent_approvals:
        event = Event.query.get(approval.event_id)
        if approval.status == 'approved':
            notifications.append({
                'type': 'success',
                'icon': 'ph-check-circle',
                'message': f'Your event "{event.title}" was approved by {approval.approver_role}!',
                'event_id': event.event_id,
                'time': approval.approved_at
            })
        elif approval.status == 'rejected':
            notifications.append({
                'type': 'error',
                'icon': 'ph-x-circle',
                'message': f'Your event "{event.title}" was rejected by {approval.approver_role}.',
                'event_id': event.event_id,
                'time': approval.approved_at,
                'remarks': approval.remarks
            })
    
    # Check for events that are now fully approved and ready to go
    newly_approved_events = Event.query.filter(
        Event.organizer_id == organizer_id,
        Event.status == 'approved',
        Event.date >= date.today()
    ).all()
    
    for ev in newly_approved_events:
        # Check if this is a recent approval (within 7 days)
        latest_approval = Approval.query.filter_by(
            event_id=ev.event_id,
            status='approved'
        ).order_by(Approval.approved_at.desc()).first()
        
        if latest_approval and latest_approval.approved_at and latest_approval.approved_at >= week_ago:
            # Check if all required approvals are done
            all_approvals = Approval.query.filter_by(event_id=ev.event_id).all()
            if all(a.status == 'approved' for a in all_approvals):
                # Only add if not already added
                if not any(n.get('event_id') == ev.event_id and n.get('type') == 'ready' for n in notifications):
                    notifications.append({
                        'type': 'ready',
                        'icon': 'ph-rocket-launch',
                        'message': f'"{ev.title}" is fully approved and ready for {ev.date.strftime("%b %d")}!',
                        'event_id': ev.event_id,
                        'time': latest_approval.approved_at
                    })
    
    # Sort notifications by time (most recent first) and limit
    notifications.sort(key=lambda x: x.get('time') or now, reverse=True)
    notifications = notifications[:5]
    
    return render_template('organizer/dashboard.html',
                         events=events,
                         pagination=pagination,
                         pending_count=pending_count,
                         approved_count=approved_count,
                         rejected_count=rejected_count,
                         past_event_ids=past_event_ids,
                         search=search,
                         notifications=notifications)


@bp.route('/create-event', methods=['GET', 'POST'])
@organizer_required
def create_event():
    """Create new event"""
    if request.method == 'POST':
        title = request.form.get('title')
        description = request.form.get('description')
        event_date = datetime.strptime(request.form.get('date'), '%Y-%m-%d').date()
        start_time = datetime.strptime(request.form.get('start_time'), '%H:%M').time()
        end_time = datetime.strptime(request.form.get('end_time'), '%H:%M').time()
        venue_id = request.form.get('venue_id')
        mode = request.form.get('mode') or 'offline'
        meeting_url = request.form.get('meeting_url')
        dept_id = request.form.get('dept_id')
        poster_file = request.files.get('poster')
        certificate_template_id = request.form.get('certificate_template_id')
        
        # Validation
        if event_date < date.today():
            flash('Event date cannot be in the past', 'error')
            return redirect(url_for('organizer.create_event'))
        
        if start_time >= end_time:
            flash('End time must be after start time', 'error')
            return redirect(url_for('organizer.create_event'))
        
        # Create event
        # If online, venue may be empty and meeting_url required
        if (mode or '').lower() == 'online':
            if not meeting_url:
                flash('Meeting URL is required for online events', 'error')
                return redirect(url_for('organizer.create_event'))
            # treat '0' or missing as no venue
            venue_val = int(venue_id) if venue_id and int(venue_id) > 0 else None
        else:
            # offline: venue required
            if not venue_id or int(venue_id) <= 0:
                flash('Venue is required for offline events', 'error')
                return redirect(url_for('organizer.create_event'))
            venue_val = int(venue_id)

        # If offline and venue selected, ensure the venue is not already booked
        if (mode or '').lower() != 'online' and venue_val is not None:
            conflict = Event.query.filter(
                Event.venue_id == venue_val,
                Event.date == event_date,
                Event.status.in_(['pending', 'approved']),
                Event.start_time < end_time,
                Event.end_time > start_time
            ).first()
            if conflict:
                flash('Selected venue is already booked for the chosen date/time', 'error')
                return redirect(url_for('organizer.create_event'))

        poster_url = None
        if poster_file and poster_file.filename:
            filename = secure_filename(poster_file.filename)
            ext = os.path.splitext(filename)[1].lower()
            allowed_exts = {'.jpg', '.jpeg', '.png', '.webp'}
            if ext not in allowed_exts:
                flash('Poster must be a JPG, PNG, or WEBP image', 'error')
                return redirect(url_for('organizer.create_event'))

            safe_name = f"event_{session['user_id']}_{int(datetime.utcnow().timestamp())}{ext}"
            poster_dir = os.path.join(current_app.config['UPLOAD_FOLDER'], 'events')
            os.makedirs(poster_dir, exist_ok=True)
            poster_path = os.path.join(poster_dir, safe_name)
            poster_file.save(poster_path)
            poster_url = os.path.join('uploads', 'events', safe_name)

        event = Event(
            title=title,
            description=description,
            date=event_date,
            start_time=start_time,
            end_time=end_time,
            venue_id=venue_val,
            dept_id=int(dept_id) if dept_id else None,
            organizer_id=session['user_id'],
            status='pending',
            mode=(mode or 'offline'),
            meeting_url=meeting_url if meeting_url else None,
            poster_url=poster_url,
            scan_token=uuid.uuid4().hex,
            certificate_template_id=int(certificate_template_id) if certificate_template_id else None,
            is_team_event=request.form.get('is_team_event') == '1',
            min_team_size=int(request.form.get('min_team_size', 2)) if request.form.get('is_team_event') == '1' else 1,
            max_team_size=int(request.form.get('max_team_size', 4)) if request.form.get('is_team_event') == '1' else 1,
            has_prizes=request.form.get('has_prizes') == '1',
            duty_leave_provided=request.form.get('duty_leave') == '1'
        )
        # Audience setting: campus-exclusive or public
        audience_val = request.form.get('audience', 'public')
        event.is_campus_exclusive = (audience_val == 'campus')
        
        # Validate venue FK before committing: ensure it references an existing venue
        if event.venue_id is not None:
            existing_venue = Venue.query.get(event.venue_id)
            if not existing_venue:
                # If mode is online, allow missing venue by setting NULL; otherwise reject
                if (event.mode or '').lower() == 'online':
                    event.venue_id = None
                else:
                    flash('Selected venue was not found. Please pick a valid venue.', 'error')
                    return redirect(url_for('organizer.create_event'))

        db.session.add(event)
        try:
            db.session.commit()
        except Exception:
            db.session.rollback()
            flash('Failed to create event due to database error.', 'error')
            return redirect(url_for('organizer.create_event'))
        
        # Create approval workflow
        # Determine department to check for HOD approval: prefer venue.dept_id if venue provided,
        # otherwise use the dept selected on the form (useful for online events tied to a dept)
        venue = Venue.query.get(venue_val) if venue_val else None
        dept_to_check = None
        if venue and venue.dept_id:
            dept_to_check = venue.dept_id
        else:
            try:
                dept_to_check = int(dept_id) if dept_id else None
            except Exception:
                dept_to_check = None

        # If there's a department to check, find HOD and create HOD approval
        hod_required = False
        if dept_to_check:
            hod = User.query.join(User.role).filter(
                User.dept_id == dept_to_check,
                User.role.has(role_name='HOD')
            ).first()
            if hod:
                hod_required = True
                approval = Approval(
                    event_id=event.event_id,
                    approver_id=hod.user_id,
                    approver_role='HOD'
                )
                db.session.add(approval)

                # Notify HOD via email
                if hod.email:
                                        base_url = os.getenv('APP_BASE_URL') or request.url_root.rstrip('/')
                                        login_url = f"{base_url}{url_for('auth.login')}"
                                        subject = f"New event awaiting approval: {event.title}"
                                        body = (
                                                f"Hello {hod.full_name},\n\n"
                                                f"A new event requires your approval.\n\n"
                                                f"Event: {event.title}\n"
                                                f"Organizer: {event.organizer.full_name if event.organizer else 'N/A'}\n"
                                                f"Date: {event.date.strftime('%Y-%m-%d')}\n"
                                                f"Time: {event.start_time.strftime('%H:%M')} - {event.end_time.strftime('%H:%M')}\n"
                                                f"Venue: {event.venue.venue_name if event.venue else 'N/A'}\n\n"
                                                f"Login: {login_url}\n"
                                        )
                                        html_body = f"""
                                        <div style=\"font-family: Arial, sans-serif; line-height:1.6; color:#111827;\">
                                            <h2 style=\"margin:0 0 12px;\">New Event Approval</h2>
                                            <p>Hello {hod.full_name},</p>
                                            <p>A new event requires your approval.</p>
                                            <table style=\"border-collapse:collapse; margin:12px 0;\">
                                                <tr><td style=\"padding:4px 8px; font-weight:600;\">Event</td><td style=\"padding:4px 8px;\">{event.title}</td></tr>
                                                <tr><td style=\"padding:4px 8px; font-weight:600;\">Organizer</td><td style=\"padding:4px 8px;\">{event.organizer.full_name if event.organizer else 'N/A'}</td></tr>
                                                <tr><td style=\"padding:4px 8px; font-weight:600;\">Date</td><td style=\"padding:4px 8px;\">{event.date.strftime('%Y-%m-%d')}</td></tr>
                                                <tr><td style=\"padding:4px 8px; font-weight:600;\">Time</td><td style=\"padding:4px 8px;\">{event.start_time.strftime('%H:%M')} - {event.end_time.strftime('%H:%M')}</td></tr>
                                                <tr><td style=\"padding:4px 8px; font-weight:600;\">Venue</td><td style=\"padding:4px 8px;\">{event.venue.venue_name if event.venue else 'N/A'}</td></tr>
                                            </table>
                                            <p>
                                                <a href=\"{login_url}\" style=\"display:inline-block; background:#4f46e5; color:#fff; text-decoration:none; padding:10px 16px; border-radius:8px;\">Open Dashboard</a>
                                            </p>
                                            <p style=\"font-size:12px; color:#6b7280;\">If the button doesn’t work, copy this link: {login_url}</p>
                                        </div>
                                        """
                                        try:
                                                send_email(hod.email, subject, body, html_body)
                                        except Exception as exc:
                                                current_app.logger.warning(f"HOD email send failed: {exc}")
                                                flash('Event created, but email notification failed. Check SMTP settings.', 'warning')
        
        # Always create Principal approval entry (will be pending until HOD approves if needed)
        principal = User.query.join(User.role).filter(
            User.role.has(role_name='Principal')
        ).first()
        
        if principal:
            principal_approval = Approval(
                event_id=event.event_id,
                approver_id=principal.user_id,
                approver_role='Principal'
            )
            db.session.add(principal_approval)
            # Notify Principal via email only when no HOD approval is required
            if not hod_required and principal.email:
                base_url = os.getenv('APP_BASE_URL') or request.url_root.rstrip('/')
                login_url = f"{base_url}{url_for('auth.login')}"
                subject = f"New event awaiting approval: {event.title}"
                body = (
                    f"Hello {principal.full_name},\n\n"
                    "A new event requires your approval.\n\n"
                    f"Event: {event.title}\n"
                    f"Organizer: {event.organizer.full_name if event.organizer else 'N/A'}\n"
                    f"Date: {event.date.strftime('%Y-%m-%d')}\n"
                    f"Time: {event.start_time.strftime('%H:%M')} - {event.end_time.strftime('%H:%M')}\n"
                    f"Venue: {event.venue.venue_name if event.venue else 'N/A'}\n\n"
                    f"Login: {login_url}\n"
                )
                try:
                    send_email(principal.email, subject, body)
                except Exception as exc:
                    current_app.logger.warning(f"Principal email send failed: {exc}")
        
        db.session.commit()
        
        flash('Event created and submitted for approval!', 'success')
        return redirect(url_for('organizer.dashboard'))
    
    # Get venues, departments, and templates
    venues = Venue.query.all()
    departments = Department.query.all()
    templates = CertificateTemplate.query.filter_by(organizer_id=session['user_id']).order_by(
        CertificateTemplate.created_at.desc()
    ).all()
    
    return render_template('organizer/create_event.html',
                         venues=venues,
                         departments=departments,
                         templates=templates)


@bp.route('/event/<int:event_id>/edit', methods=['GET', 'POST'])
@organizer_required
def edit_event(event_id):
    """Edit an existing event"""
    organizer_id = session['user_id']
    event = Event.query.filter_by(event_id=event_id, organizer_id=organizer_id).first_or_404()

    if request.method == 'POST':
        title = request.form.get('title')
        description = request.form.get('description')
        event_date = datetime.strptime(request.form.get('date'), '%Y-%m-%d').date()
        start_time = datetime.strptime(request.form.get('start_time'), '%H:%M').time()
        end_time = datetime.strptime(request.form.get('end_time'), '%H:%M').time()
        venue_id = request.form.get('venue_id')
        mode = request.form.get('mode') or 'offline'
        meeting_url = request.form.get('meeting_url')
        dept_id = request.form.get('dept_id')
        certificate_template_id = request.form.get('certificate_template_id')

        if event_date < date.today():
            flash('Event date cannot be in the past', 'error')
            return redirect(url_for('organizer.edit_event', event_id=event_id))

        if start_time >= end_time:
            flash('End time must be after start time', 'error')
            return redirect(url_for('organizer.edit_event', event_id=event_id))

        if (mode or '').lower() == 'online':
            if not meeting_url:
                flash('Meeting URL is required for online events', 'error')
                return redirect(url_for('organizer.edit_event', event_id=event_id))
            venue_val = int(venue_id) if venue_id and int(venue_id) > 0 else None
        else:
            if not venue_id or int(venue_id) <= 0:
                flash('Venue is required for offline events', 'error')
                return redirect(url_for('organizer.edit_event', event_id=event_id))
            venue_val = int(venue_id)

        if (mode or '').lower() != 'online' and venue_val is not None:
            conflict = Event.query.filter(
                Event.venue_id == venue_val,
                Event.date == event_date,
                Event.status.in_(['pending', 'approved']),
                Event.start_time < end_time,
                Event.end_time > start_time,
                Event.event_id != event.event_id
            ).first()
            if conflict:
                flash('Selected venue is already booked for the chosen date/time', 'error')
                return redirect(url_for('organizer.edit_event', event_id=event_id))

        requires_reapproval = (
            event.date != event_date or
            event.start_time != start_time or
            event.end_time != end_time or
            (event.venue_id or None) != venue_val or
            (event.mode or 'offline') != (mode or 'offline') or
            (event.meeting_url or None) != (meeting_url or None) or
            (event.dept_id or None) != (int(dept_id) if dept_id else None)
        )

        event.title = title
        event.description = description
        event.date = event_date
        event.start_time = start_time
        event.end_time = end_time
        event.venue_id = venue_val
        event.dept_id = int(dept_id) if dept_id else None
        event.mode = (mode or 'offline')
        event.meeting_url = meeting_url if meeting_url else None
        event.certificate_template_id = int(certificate_template_id) if certificate_template_id else None
        event.has_prizes = request.form.get('has_prizes') == '1'
        event.duty_leave_provided = request.form.get('duty_leave') == '1'

        if requires_reapproval:
            event.status = 'pending'
            Approval.query.filter_by(event_id=event.event_id).delete()

            venue = Venue.query.get(venue_val) if venue_val else None
            dept_to_check = venue.dept_id if venue and venue.dept_id else event.dept_id
            if dept_to_check:
                hod = User.query.join(User.role).filter(
                    User.dept_id == dept_to_check,
                    User.role.has(role_name='HOD')
                ).first()
                if hod:
                    db.session.add(Approval(
                        event_id=event.event_id,
                        approver_id=hod.user_id,
                        approver_role='HOD',
                        status='pending',
                        remarks='Event updated (not new). Approval required.'
                    ))

            principal = User.query.join(User.role).filter(
                User.role.has(role_name='Principal')
            ).first()
            if principal:
                db.session.add(Approval(
                    event_id=event.event_id,
                    approver_id=principal.user_id,
                    approver_role='Principal',
                    status='pending',
                    remarks='Event updated (not new). Approval required.'
                ))

            flash('Event updated and re-submitted for approval (not a new event).', 'success')
        else:
            flash('Event updated successfully.', 'success')

        db.session.commit()
        return redirect(url_for('organizer.view_event', event_id=event.event_id))

    venues = Venue.query.all()
    departments = Department.query.all()
    templates = CertificateTemplate.query.filter_by(organizer_id=organizer_id).order_by(
        CertificateTemplate.created_at.desc()
    ).all()
    return render_template('organizer/edit_event.html', event=event, venues=venues, departments=departments, templates=templates)


@bp.route('/certificate-templates')
@organizer_required
def certificate_templates():
    """List organizer certificate templates"""
    organizer_id = session['user_id']
    templates = CertificateTemplate.query.filter_by(organizer_id=organizer_id).order_by(
        CertificateTemplate.created_at.desc()
    ).all()
    return render_template('organizer/certificate_templates.html', templates=templates, max_templates=3)


@bp.route('/certificate-templates/upload', methods=['POST'])
@organizer_required
def upload_certificate_template():
    """Upload a new certificate template (max 10 per organizer)"""
    organizer_id = session['user_id']
    existing_count = CertificateTemplate.query.filter_by(organizer_id=organizer_id).count()
    if existing_count >= 10:
        flash('You can upload up to 10 certificate templates only.', 'error')
        return redirect(url_for('organizer.certificate_templates'))

    name = (request.form.get('name') or '').strip() or f'Template {existing_count + 1}'
    template_file = request.files.get('template_image')

    if not template_file or not template_file.filename:
        flash('Please select a template image to upload.', 'error')
        return redirect(url_for('organizer.certificate_templates'))

    filename = secure_filename(template_file.filename)
    ext = os.path.splitext(filename)[1].lower()
    allowed_exts = {'.jpg', '.jpeg', '.png', '.webp'}
    if ext not in allowed_exts:
        flash('Template must be a JPG, PNG, or WEBP image.', 'error')
        return redirect(url_for('organizer.certificate_templates'))

    safe_name = f"cert_template_{organizer_id}_{int(datetime.utcnow().timestamp())}{ext}"
    template_dir = os.path.join(current_app.config['UPLOAD_FOLDER'], 'certificates', 'templates')
    os.makedirs(template_dir, exist_ok=True)
    template_path = os.path.join(template_dir, safe_name)
    template_file.save(template_path)

    is_default = existing_count == 0
    template = CertificateTemplate(
        organizer_id=organizer_id,
        name=name,
        image_url=os.path.join('uploads', 'certificates', 'templates', safe_name),
        is_default=is_default
    )
    db.session.add(template)
    db.session.commit()

    flash('Template uploaded successfully.', 'success')
    return redirect(url_for('organizer.certificate_templates'))


@bp.route('/certificate-templates/<int:template_id>/edit', methods=['GET', 'POST'])
@organizer_required
def edit_certificate_template(template_id):
    """Interactive editor to set certificate text positions"""
    organizer_id = session['user_id']
    template = CertificateTemplate.query.filter_by(
        template_id=template_id,
        organizer_id=organizer_id
    ).first_or_404()

    if request.method == 'POST':
        positions_json = (request.form.get('positions_json') or '').strip()
        if positions_json:
            try:
                json.loads(positions_json)
                template.positions = positions_json
            except Exception:
                flash('Invalid positions data. Please try again.', 'error')
                return redirect(url_for('organizer.edit_certificate_template', template_id=template_id))
        else:
            template.positions = None

        db.session.commit()
        flash('Template positions saved.', 'success')
        return redirect(url_for('organizer.edit_certificate_template', template_id=template_id, preview='1'))

    positions = {}
    if template.positions:
        try:
            positions = json.loads(template.positions)
        except Exception:
            positions = {}

    show_preview = (request.args.get('preview') == '1')
    return render_template(
        'organizer/certificate_template_edit.html',
        template=template,
        positions=positions,
        show_preview=show_preview
    )


@bp.route('/certificate-templates/<int:template_id>/set-default', methods=['POST'])
@organizer_required
def set_default_certificate_template(template_id):
    """Set a default certificate template for the organizer"""
    organizer_id = session['user_id']
    template = CertificateTemplate.query.filter_by(
        template_id=template_id,
        organizer_id=organizer_id
    ).first_or_404()

    CertificateTemplate.query.filter_by(organizer_id=organizer_id).update({'is_default': False})
    template.is_default = True
    db.session.commit()

    flash('Default certificate template updated.', 'success')
    return redirect(url_for('organizer.certificate_templates'))


@bp.route('/certificate-templates/<int:template_id>/delete', methods=['POST'])
@organizer_required
def delete_certificate_template(template_id):
    """Delete a certificate template"""
    organizer_id = session['user_id']
    template = CertificateTemplate.query.filter_by(
        template_id=template_id,
        organizer_id=organizer_id
    ).first_or_404()

    template_path = os.path.join('static', template.image_url)
    was_default = template.is_default

    db.session.delete(template)
    db.session.commit()

    if os.path.exists(template_path):
        try:
            os.remove(template_path)
        except Exception:
            pass

    if was_default:
        next_template = CertificateTemplate.query.filter_by(organizer_id=organizer_id).order_by(
            CertificateTemplate.created_at.desc()
        ).first()
        if next_template:
            next_template.is_default = True
            db.session.commit()

    flash('Template deleted.', 'success')
    return redirect(url_for('organizer.certificate_templates'))


@bp.route('/event/<int:event_id>/delete', methods=['POST'])
@organizer_required
def delete_event(event_id):
    """Delete an event"""
    organizer_id = session['user_id']
    event = Event.query.filter_by(event_id=event_id, organizer_id=organizer_id).first_or_404()
    db.session.delete(event)
    db.session.commit()
    flash('Event deleted successfully.', 'success')
    return redirect(url_for('organizer.dashboard'))


@bp.route('/event/<int:event_id>')
@organizer_required
def view_event(event_id):
    """View event details and registrations"""
    from models.models import Team, CertificateTemplate
    organizer_id = session['user_id']
    
    event = Event.query.filter_by(
        event_id=event_id,
        organizer_id=organizer_id
    ).first_or_404()
    
    # Get registrations
    registrations = Registration.query.filter_by(event_id=event_id).all()
    
    # Get attendance count
    attended_count = db.session.query(Attendance).join(Registration).filter(
        Registration.event_id == event_id
    ).count()
    
    # Get approvals
    approvals = Approval.query.filter_by(event_id=event_id).order_by(
        Approval.approved_at
    ).all()
    
    # Get teams for team events
    teams = []
    if event.is_team_event:
        teams = Team.query.filter_by(event_id=event_id).all()
    
    # Get certificate templates for prize assignment
    certificate_templates = CertificateTemplate.query.filter_by(organizer_id=organizer_id).all()
    
    # Check if event is past (for showing Reviews button)
    past_event_ids = set()
    now = datetime.now()
    try:
        end_dt = datetime.combine(event.date, event.end_time)
        if now >= end_dt:
            past_event_ids.add(event.event_id)
    except Exception:
        pass
    
    return render_template('organizer/view_event.html',
                         event=event,
                         registrations=registrations,
                         attended_count=attended_count,
                         approvals=approvals,
                         past_event_ids=past_event_ids,
                         teams=teams,
                         certificate_templates=certificate_templates)


@bp.route('/event/<int:event_id>/assign-prize', methods=['POST'])
@organizer_required
def assign_prize(event_id):
    """Assign prize to a team and regenerate certificates"""
    from models.models import Team
    organizer_id = session['user_id']
    
    # Verify event belongs to organizer
    event = Event.query.filter_by(
        event_id=event_id,
        organizer_id=organizer_id
    ).first_or_404()
    
    if not event.is_team_event:
        flash('Use the individual prize assignment for non-team events.', 'error')
        return redirect(url_for('organizer.view_event', event_id=event_id))
    
    team_id = request.form.get('team_id')
    prize_position = request.form.get('prize_position')
    prize_title = request.form.get('prize_title', '').strip()
    certificate_template_id = request.form.get('certificate_template_id')
    
    if not team_id or not prize_position:
        flash('Please select a team and prize position.', 'error')
        return redirect(url_for('organizer.view_event', event_id=event_id))
    
    # Get team
    team = Team.query.filter_by(team_id=team_id, event_id=event_id).first()
    if not team:
        flash('Team not found.', 'error')
        return redirect(url_for('organizer.view_event', event_id=event_id))
    
    # Check if minimum required team members have attendance
    attended_members = [m for m in team.members if m.attendance]
    min_required = event.min_team_size
    if len(attended_members) < min_required:
        flash(f'Cannot assign prize. At least {min_required} team members must have attended (currently {len(attended_members)} attended).', 'error')
        return redirect(url_for('organizer.view_event', event_id=event_id))
    
    # Update prize info
    team.prize_position = prize_position
    team.prize_title = prize_title if prize_title else None
    team.prize_certificate_template_id = int(certificate_template_id) if certificate_template_id else None
    
    db.session.commit()
    
    # Auto-regenerate certificates for team members who have attendance
    regenerated_count = 0
    for member in team.members:
        if member.attendance:
            # Delete existing certificate
            existing_cert = Certificate.query.filter_by(
                student_id=member.student_id,
                event_id=event_id
            ).first()
            
            if existing_cert:
                # Delete the file if it exists
                cert_file_path = os.path.join('static', existing_cert.certificate_url)
                if os.path.exists(cert_file_path):
                    os.remove(cert_file_path)
                db.session.delete(existing_cert)
                db.session.commit()
            
            # Generate new certificate with prize info
            generate_certificate_for_student(member.student_id, event_id)
            regenerated_count += 1
    
    if regenerated_count > 0:
        flash(f'Prize "{prize_position}" assigned to team "{team.team_name}"! {regenerated_count} certificate(s) updated.', 'success')
    else:
        flash(f'Prize "{prize_position}" assigned to team "{team.team_name}"!', 'success')
    
    return redirect(url_for('organizer.view_event', event_id=event_id))


@bp.route('/event/<int:event_id>/clear-prize', methods=['POST'])
@organizer_required
def clear_prize(event_id):
    """Clear prize from a team and regenerate certificates"""
    from models.models import Team
    organizer_id = session['user_id']
    
    # Verify event belongs to organizer
    event = Event.query.filter_by(
        event_id=event_id,
        organizer_id=organizer_id
    ).first_or_404()
    
    team_id = request.form.get('team_id')
    
    if not team_id:
        flash('Please select a team.', 'error')
        return redirect(url_for('organizer.view_event', event_id=event_id))
    
    # Get team
    team = Team.query.filter_by(team_id=team_id, event_id=event_id).first()
    if not team:
        flash('Team not found.', 'error')
        return redirect(url_for('organizer.view_event', event_id=event_id))
    
    # Store team name before clearing
    team_name = team.team_name
    
    # Clear prize info
    team.prize_position = None
    team.prize_title = None
    team.prize_certificate_template_id = None
    
    db.session.commit()
    
    # Regenerate certificates (now without prize info)
    regenerated_count = 0
    for member in team.members:
        if member.attendance:
            existing_cert = Certificate.query.filter_by(
                student_id=member.student_id,
                event_id=event_id
            ).first()
            
            if existing_cert:
                cert_file_path = os.path.join('static', existing_cert.certificate_url)
                if os.path.exists(cert_file_path):
                    os.remove(cert_file_path)
                db.session.delete(existing_cert)
                db.session.commit()
            
            generate_certificate_for_student(member.student_id, event_id)
            regenerated_count += 1
    
    if regenerated_count > 0:
        flash(f'Prize cleared from team "{team_name}". {regenerated_count} certificate(s) updated.', 'success')
    else:
        flash(f'Prize cleared from team "{team_name}".', 'success')
    
    return redirect(url_for('organizer.view_event', event_id=event_id))


@bp.route('/event/<int:event_id>/assign-individual-prize', methods=['POST'])
@organizer_required
def assign_individual_prize(event_id):
    """Assign prize to an individual participant and regenerate certificate"""
    organizer_id = session['user_id']
    
    # Verify event belongs to organizer
    event = Event.query.filter_by(
        event_id=event_id,
        organizer_id=organizer_id
    ).first_or_404()
    
    if event.is_team_event:
        flash('Use team prize assignment for team events.', 'error')
        return redirect(url_for('organizer.view_event', event_id=event_id))
    
    if not event.has_prizes:
        flash('This event does not have prizes enabled.', 'error')
        return redirect(url_for('organizer.view_event', event_id=event_id))
    
    registration_id = request.form.get('registration_id')
    prize_position = request.form.get('prize_position')
    prize_title = request.form.get('prize_title', '').strip()
    certificate_template_id = request.form.get('certificate_template_id')
    
    if not registration_id or not prize_position:
        flash('Please select a participant and prize position.', 'error')
        return redirect(url_for('organizer.view_event', event_id=event_id))
    
    # Get registration
    registration = Registration.query.filter_by(
        registration_id=registration_id, 
        event_id=event_id
    ).first()
    
    if not registration:
        flash('Participant not found.', 'error')
        return redirect(url_for('organizer.view_event', event_id=event_id))
    
    # Check if participant has attendance
    if not registration.attendance:
        flash('Cannot assign prize to a participant who has not attended the event.', 'error')
        return redirect(url_for('organizer.view_event', event_id=event_id))
    
    # Update prize info
    registration.prize_position = prize_position
    registration.prize_title = prize_title if prize_title else None
    registration.prize_certificate_template_id = int(certificate_template_id) if certificate_template_id else None
    
    db.session.commit()
    
    # Regenerate certificate with prize info
    existing_cert = Certificate.query.filter_by(
        student_id=registration.student_id,
        event_id=event_id
    ).first()
    
    if existing_cert:
        cert_file_path = os.path.join('static', existing_cert.certificate_url)
        if os.path.exists(cert_file_path):
            os.remove(cert_file_path)
        db.session.delete(existing_cert)
        db.session.commit()
    
    generate_certificate_for_student(registration.student_id, event_id)
    flash(f'Prize "{prize_position}" assigned to {registration.student.full_name}! Certificate updated.', 'success')
    
    return redirect(url_for('organizer.view_event', event_id=event_id))


@bp.route('/event/<int:event_id>/clear-individual-prize', methods=['POST'])
@organizer_required
def clear_individual_prize(event_id):
    """Clear prize from an individual participant and regenerate certificate"""
    organizer_id = session['user_id']
    
    # Verify event belongs to organizer
    event = Event.query.filter_by(
        event_id=event_id,
        organizer_id=organizer_id
    ).first_or_404()
    
    registration_id = request.form.get('registration_id')
    
    if not registration_id:
        flash('Please select a participant.', 'error')
        return redirect(url_for('organizer.view_event', event_id=event_id))
    
    # Get registration
    registration = Registration.query.filter_by(
        registration_id=registration_id, 
        event_id=event_id
    ).first()
    
    if not registration:
        flash('Participant not found.', 'error')
        return redirect(url_for('organizer.view_event', event_id=event_id))
    
    # Store name before clearing
    student_name = registration.student.full_name
    
    # Clear prize info
    registration.prize_position = None
    registration.prize_title = None
    
    db.session.commit()
    
    # Regenerate certificate (now without prize info)
    if registration.attendance:
        existing_cert = Certificate.query.filter_by(
            student_id=registration.student_id,
            event_id=event_id
        ).first()
        
        if existing_cert:
            cert_file_path = os.path.join('static', existing_cert.certificate_url)
            if os.path.exists(cert_file_path):
                os.remove(cert_file_path)
            db.session.delete(existing_cert)
            db.session.commit()
        
        generate_certificate_for_student(registration.student_id, event_id)
        flash(f'Prize cleared from {student_name}. Certificate updated.', 'success')
    else:
        flash(f'Prize cleared from {student_name}.', 'success')
    
    return redirect(url_for('organizer.view_event', event_id=event_id))


@bp.route('/event/<int:event_id>/regenerate-team-certificates', methods=['POST'])
@organizer_required
def regenerate_team_certificates(event_id):
    """Regenerate certificates for all team members (for prize updates)"""
    from models.models import Team
    organizer_id = session['user_id']
    
    # Verify event belongs to organizer
    event = Event.query.filter_by(
        event_id=event_id,
        organizer_id=organizer_id
    ).first_or_404()
    
    if not event.is_team_event:
        flash('This is not a team event.', 'error')
        return redirect(url_for('organizer.view_event', event_id=event_id))
    
    team_id = request.form.get('team_id')
    
    if not team_id:
        flash('Please select a team.', 'error')
        return redirect(url_for('organizer.view_event', event_id=event_id))
    
    # Get team
    team = Team.query.filter_by(team_id=team_id, event_id=event_id).first()
    if not team:
        flash('Team not found.', 'error')
        return redirect(url_for('organizer.view_event', event_id=event_id))
    
    regenerated_count = 0
    
    # Regenerate certificates for all team members who have attendance
    for member in team.members:
        if member.attendance:
            # Delete existing certificate
            existing_cert = Certificate.query.filter_by(
                student_id=member.student_id,
                event_id=event_id
            ).first()
            
            if existing_cert:
                # Delete the file if it exists
                cert_file_path = os.path.join('static', existing_cert.certificate_url)
                if os.path.exists(cert_file_path):
                    os.remove(cert_file_path)
                db.session.delete(existing_cert)
                db.session.commit()
            
            # Generate new certificate with prize info
            generate_certificate_for_student(member.student_id, event_id)
            regenerated_count += 1
    
    if regenerated_count > 0:
        flash(f'Regenerated {regenerated_count} certificate(s) for team "{team.team_name}".', 'success')
    else:
        flash(f'No certificates to regenerate for team "{team.team_name}" (no attendance marked yet).', 'info')
    
    return redirect(url_for('organizer.view_event', event_id=event_id))


@bp.route('/event/<int:event_id>/download-attendance/<format>')
@organizer_required
def download_attendance(event_id, format):
    """Download attendance list as Excel or PDF"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    
    organizer_id = session['user_id']
    
    event = Event.query.filter_by(
        event_id=event_id,
        organizer_id=organizer_id
    ).first_or_404()
    
    # Get attended registrations
    attended_regs = db.session.query(Registration).join(Attendance).filter(
        Registration.event_id == event_id
    ).all()
    
    # Check if team event
    is_team = event.is_team_event
    
    if format == 'excel':
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Attendance"
        
        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Title
        merge_cols = 'A1:E1' if is_team else 'A1:D1'
        ws.merge_cells(merge_cols)
        ws['A1'] = f"Attendance Report - {event.title}"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal="center")
        
        # Event details
        ws['A2'] = f"Date: {event.date.strftime('%B %d, %Y')}"
        ws['A3'] = f"Total Attended: {len(attended_regs)}"
        if is_team:
            ws['A4'] = f"Event Type: Team Event ({event.min_team_size}-{event.max_team_size} members)"
        
        # Headers
        if is_team:
            headers = ['S.No', 'Team Name', 'Student Name', 'Username', 'Email']
        else:
            headers = ['S.No', 'Student Name', 'Username', 'Email']
        
        start_row = 6 if is_team else 5
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Data rows
        for idx, reg in enumerate(attended_regs, 1):
            row = idx + start_row
            if is_team:
                team_name = reg.team.team_name if reg.team else '-'
                ws.cell(row=row, column=1, value=idx).border = thin_border
                ws.cell(row=row, column=2, value=team_name).border = thin_border
                ws.cell(row=row, column=3, value=reg.student.full_name).border = thin_border
                ws.cell(row=row, column=4, value=reg.student.username).border = thin_border
                ws.cell(row=row, column=5, value=reg.student.email).border = thin_border
            else:
                ws.cell(row=row, column=1, value=idx).border = thin_border
                ws.cell(row=row, column=2, value=reg.student.full_name).border = thin_border
                ws.cell(row=row, column=3, value=reg.student.username).border = thin_border
                ws.cell(row=row, column=4, value=reg.student.email).border = thin_border
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 8
        if is_team:
            ws.column_dimensions['B'].width = 25
            ws.column_dimensions['C'].width = 25
            ws.column_dimensions['D'].width = 18
            ws.column_dimensions['E'].width = 30
        else:
            ws.column_dimensions['B'].width = 30
            ws.column_dimensions['C'].width = 20
            ws.column_dimensions['D'].width = 35
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"attendance_{event.title.replace(' ', '_')}_{event.date.strftime('%Y%m%d')}.xlsx"
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    
    elif format == 'pdf':
        # Create PDF
        output = BytesIO()
        doc = SimpleDocTemplate(output, pagesize=A4, topMargin=30, bottomMargin=30)
        elements = []
        styles = getSampleStyleSheet()
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            alignment=1,
            spaceAfter=20
        )
        elements.append(Paragraph(f"Attendance Report", title_style))
        elements.append(Paragraph(f"<b>{event.title}</b>", styles['Heading2']))
        elements.append(Spacer(1, 10))
        elements.append(Paragraph(f"Date: {event.date.strftime('%B %d, %Y')}", styles['Normal']))
        elements.append(Paragraph(f"Total Attended: {len(attended_regs)}", styles['Normal']))
        if is_team:
            elements.append(Paragraph(f"Event Type: Team Event ({event.min_team_size}-{event.max_team_size} members)", styles['Normal']))
        elements.append(Spacer(1, 20))
        
        # Table data
        if is_team:
            data = [['S.No', 'Team Name', 'Student Name', 'Username', 'Email']]
            for idx, reg in enumerate(attended_regs, 1):
                team_name = reg.team.team_name if reg.team else '-'
                data.append([
                    str(idx),
                    team_name,
                    reg.student.full_name,
                    reg.student.username,
                    reg.student.email
                ])
            col_widths = [30, 90, 120, 80, 150]
        else:
            data = [['S.No', 'Student Name', 'Username', 'Email']]
            for idx, reg in enumerate(attended_regs, 1):
                data.append([
                    str(idx),
                    reg.student.full_name,
                    reg.student.username,
                    reg.student.email
                ])
            col_widths = [40, 150, 100, 180]
        
        # Create table
        table = Table(data, colWidths=col_widths)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4F46E5')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F3F4F6')]),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 1), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
        ]))
        elements.append(table)
        
        # Build PDF
        doc.build(elements)
        output.seek(0)
        
        filename = f"attendance_{event.title.replace(' ', '_')}_{event.date.strftime('%Y%m%d')}.pdf"
        return send_file(
            output,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=filename
        )
    
    flash('Invalid format specified', 'error')
    return redirect(url_for('organizer.view_event', event_id=event_id))


@bp.route('/event/<int:event_id>/attendance-upload', methods=['POST'])
@organizer_required
def attendance_upload(event_id):
    """Upload Excel attendance for online events"""
    organizer_id = session['user_id']

    event = Event.query.filter_by(
        event_id=event_id,
        organizer_id=organizer_id
    ).first_or_404()

    if (event.mode or '').lower() != 'online':
        flash('Attendance upload is available only for online events.', 'warning')
        return redirect(url_for('organizer.view_event', event_id=event.event_id))

    try:
        event_start = datetime.combine(event.date, event.start_time)
        if datetime.now() < event_start:
            flash('Cannot upload attendance before the event start time.', 'error')
            return redirect(url_for('organizer.view_event', event_id=event.event_id))
        event_end = datetime.combine(event.date, event.end_time)
        if datetime.now() > (event_end + timedelta(days=3)):
            flash('Attendance upload window has expired (3 days).', 'error')
            return redirect(url_for('organizer.view_event', event_id=event.event_id))
    except Exception:
        pass

    file = request.files.get('attendance_file')
    if not file or not file.filename:
        flash('Please upload an Excel file.', 'error')
        return redirect(url_for('organizer.view_event', event_id=event.event_id))

    filename = secure_filename(file.filename)
    if not filename.lower().endswith(('.xlsx', '.xlsm', '.xltx', '.xltm', '.xls')):
        flash('Invalid file type. Upload an Excel file (.xlsx).', 'error')
        return redirect(url_for('organizer.view_event', event_id=event.event_id))

    try:
        workbook = load_workbook(file, data_only=True)
        sheet = workbook.active
    except Exception:
        flash('Unable to read the Excel file. Please re-export and try again.', 'error')
        return redirect(url_for('organizer.view_event', event_id=event.event_id))

    def normalize_header(value):
        text = (value or '').strip().lower()
        return ''.join(ch for ch in text if ch.isalnum())

    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        flash('The uploaded file is empty.', 'error')
        return redirect(url_for('organizer.view_event', event_id=event.event_id))

    headers = [normalize_header(col) for col in rows[0]]
    header_map = {name: idx for idx, name in enumerate(headers) if name}

    email_keys = {'email', 'emailaddress', 'studentemail', 'useremail', 'emailid', 'mail'}
    username_keys = {'username', 'userid', 'rollno', 'rollnumber', 'studentid', 'registerno', 'regno'}

    email_idx = next((header_map[k] for k in email_keys if k in header_map), None)
    username_idx = next((header_map[k] for k in username_keys if k in header_map), None)

    if email_idx is None and username_idx is None:
        flash('No Email or Username column found. Please include Email/Username in the Excel file.', 'error')
        return redirect(url_for('organizer.view_event', event_id=event.event_id))

    registrations = Registration.query.filter_by(event_id=event_id).all()
    email_to_reg = {}
    username_to_reg = {}
    for reg in registrations:
        if reg.student and reg.student.email:
            email_to_reg[reg.student.email.strip().lower()] = reg
        if reg.student and reg.student.username:
            username_to_reg[reg.student.username.strip().lower()] = reg

    marked_count = 0
    already_count = 0
    not_registered_count = 0
    invalid_count = 0
    seen_registrations = set()
    new_attendance_regs = []

    for row in rows[1:]:
        email_val = row[email_idx] if email_idx is not None and email_idx < len(row) else None
        username_val = row[username_idx] if username_idx is not None and username_idx < len(row) else None

        identifier = None
        reg = None

        if email_val:
            identifier = str(email_val).strip().lower()
            reg = email_to_reg.get(identifier)

        if not reg and username_val:
            identifier = str(username_val).strip().lower()
            reg = username_to_reg.get(identifier)

        if not identifier:
            invalid_count += 1
            continue

        if not reg:
            not_registered_count += 1
            continue

        if reg.registration_id in seen_registrations:
            continue

        seen_registrations.add(reg.registration_id)

        existing = Attendance.query.filter_by(registration_id=reg.registration_id).first()
        if existing:
            already_count += 1
            continue

        attendance = Attendance(
            registration_id=reg.registration_id,
            scan_time=datetime.now(),
            scanned_by=organizer_id,
            status='present'
        )
        reg.attendance = attendance
        db.session.add(attendance)
        db.session.add(reg)
        new_attendance_regs.append(reg)
        marked_count += 1

    try:
        db.session.commit()
    except Exception:
        db.session.rollback()
        flash('Failed to upload attendance due to database error.', 'error')
        return redirect(url_for('organizer.view_event', event_id=event.event_id))

    for reg in new_attendance_regs:
        try:
            generate_certificate_for_student(reg.student_id, reg.event_id)
        except Exception:
            pass

    flash(
        f'Attendance upload complete. Marked: {marked_count}, Already marked: {already_count}, '
        f'Not registered: {not_registered_count}, Invalid rows: {invalid_count}.',
        'success'
    )
    return redirect(url_for('organizer.view_event', event_id=event.event_id))


@bp.route('/event/<int:event_id>/approval-pdf')
@organizer_required
def download_approval_pdf(event_id):
    """Download approval document PDF for an approved event."""
    organizer_id = session['user_id']
    event = Event.query.filter_by(
        event_id=event_id,
        organizer_id=organizer_id
    ).first_or_404()

    if event.status != 'approved':
        flash('Approval document is available only for approved events.', 'warning')
        return redirect(url_for('organizer.view_event', event_id=event_id))

    approvals = Approval.query.filter_by(event_id=event_id).all()
    hod_approval = next((a for a in approvals if (a.approver_role or '').lower() == 'hod'), None)
    principal_approval = next((a for a in approvals if (a.approver_role or '').lower() == 'principal'), None)

    approved_dates = [a.approved_at for a in approvals if a.status == 'approved' and a.approved_at]
    approved_on = max(approved_dates) if approved_dates else None

    buffer = BytesIO()
    pdf = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4

    left = 20 * mm
    y = height - 25 * mm

    def draw_wrapped_line(label, value, max_width):
        nonlocal y
        pdf.setFont("Helvetica", 11)
        text = f"{label}{value}" if label else str(value)
        words = text.split()
        line = ""
        for word in words:
            test_line = f"{line} {word}".strip()
            if pdf.stringWidth(test_line, "Helvetica", 11) <= max_width:
                line = test_line
            else:
                pdf.drawString(left, y, line)
                y -= 6 * mm
                line = word
        if line:
            pdf.drawString(left, y, line)
            y -= 6 * mm

    # Letterhead
    pdf.setFont("Helvetica-Bold", 16)
    pdf.drawCentredString(width / 2, y, "COLLEGE OF ENGINEERING THALASSERY")
    y -= 6 * mm
    pdf.setFont("Helvetica", 10)
    pdf.drawCentredString(width / 2, y, "Eranholi (P.O), Thalassery, Kannur, Kerala – 670107")
    y -= 5 * mm
    pdf.drawCentredString(width / 2, y, "(An Institution of Kerala Technological University)")
    y -= 6 * mm
    pdf.setLineWidth(0.8)
    pdf.line(left, y, width - left, y)
    y -= 8 * mm

    # Reference block
    ref_no = f"CET/EVENT/{event.event_id}/{datetime.now().strftime('%Y')}"
    pdf.setFont("Helvetica", 11)
    pdf.drawString(left, y, f"Ref: {ref_no}")
    pdf.drawRightString(width - left, y, f"Date: {datetime.now().strftime('%Y-%m-%d')}")
    y -= 10 * mm

    # Subject
    pdf.setFont("Helvetica-Bold", 12)
    pdf.drawString(left, y, "Subject: Approval of Event")
    y -= 8 * mm

    # Body
    pdf.setFont("Helvetica", 11)
    body_intro = (
        "This is to certify that the following event has been reviewed and approved by the "
        "competent authorities of the College of Engineering Thalassery."
    )
    draw_wrapped_line("", body_intro, width - (2 * left))
    y -= 2 * mm

    # Event details
    pdf.setFont("Helvetica-Bold", 12)
    pdf.drawString(left, y, "Event Details")
    y -= 8 * mm
    pdf.setFont("Helvetica", 11)
    pdf.drawString(left, y, f"Event Title: {event.title}")
    y -= 7 * mm
    pdf.drawString(left, y, f"Organizer: {event.organizer.full_name if event.organizer else '—'}")
    y -= 7 * mm
    pdf.drawString(left, y, f"Department: {event.department.dept_name if event.department else '—'}")
    y -= 7 * mm
    if event.description:
        draw_wrapped_line("Description: ", event.description, width - (2 * left))
    pdf.drawString(left, y, f"Date: {event.date.strftime('%Y-%m-%d')}")
    y -= 7 * mm
    pdf.drawString(left, y, f"Time: {event.start_time.strftime('%H:%M')} - {event.end_time.strftime('%H:%M')}")
    y -= 7 * mm
    pdf.drawString(left, y, f"Mode: {'Online' if (event.mode or '').lower() == 'online' else 'Offline'}")
    y -= 7 * mm
    pdf.drawString(left, y, f"Venue: {event.venue.venue_name if event.venue else '—'}")
    y -= 7 * mm
    if (event.mode or '').lower() == 'online' and event.meeting_url:
        pdf.drawString(left, y, f"Meeting Link: {event.meeting_url}")
        y -= 7 * mm

    if approved_on:
        pdf.drawString(left, y, f"Approved On: {approved_on.strftime('%Y-%m-%d %H:%M')}")
    else:
        pdf.drawString(left, y, "Approved On: —")
    y -= 10 * mm

    pdf.setFont("Helvetica-Bold", 12)
    pdf.drawString(left, y, "Approval & Remarks")
    y -= 8 * mm
    pdf.setFont("Helvetica", 11)

    def draw_approval_section(label, approval_obj):
        nonlocal y
        pdf.setFont("Helvetica-Bold", 11)
        pdf.drawString(left, y, f"{label}")
        y -= 6 * mm
        pdf.setFont("Helvetica", 11)
        name = approval_obj.approver.full_name if approval_obj and approval_obj.approver else '—'
        status = approval_obj.status if approval_obj else '—'
        remarks = approval_obj.remarks if approval_obj and approval_obj.remarks else '—'
        approved_at = approval_obj.approved_at.strftime('%Y-%m-%d %H:%M') if approval_obj and approval_obj.approved_at else '—'
        pdf.drawString(left, y, f"Name: {name}")
        y -= 6 * mm
        pdf.drawString(left, y, f"Status: {status}")
        y -= 6 * mm
        draw_wrapped_line("Remarks: ", remarks, width - (2 * left))
        pdf.drawString(left, y, f"Signed On: {approved_at}")
        y -= 10 * mm

    draw_approval_section("HOD", hod_approval)
    draw_approval_section("Principal", principal_approval)

    # Signature lines
    sig_y = 30 * mm
    pdf.setLineWidth(0.6)
    pdf.line(left, sig_y, left + 55 * mm, sig_y)
    pdf.line(width - left - 55 * mm, sig_y, width - left, sig_y)
    pdf.setFont("Helvetica", 10)
    pdf.drawString(left, sig_y - 5 * mm, "HOD Signature")
    pdf.drawRightString(width - left, sig_y - 5 * mm, "Principal Signature")

    pdf.setFont("Helvetica-Oblique", 9)
    pdf.drawString(left, 15 * mm, "This is a system-generated approval document.")

    pdf.showPage()
    pdf.save()
    buffer.seek(0)

    filename = f"approval_event_{event.event_id}_{datetime.now().strftime('%Y%m%d%H%M%S')}.pdf"
    return send_file(
        buffer,
        as_attachment=True,
        download_name=filename,
        mimetype='application/pdf'
    )


@bp.route('/event/<int:event_id>/feedback')
@organizer_required
def event_feedback(event_id):
    """View feedback for a past event"""
    organizer_id = session['user_id']
    event = Event.query.filter_by(
        event_id=event_id,
        organizer_id=organizer_id
    ).first_or_404()

    try:
        end_dt = datetime.combine(event.date, event.end_time)
        if datetime.now() < end_dt:
            flash('Feedback is available only after the event has ended.', 'warning')
            return redirect(url_for('organizer.dashboard'))
    except Exception:
        flash('Feedback is available only after the event has ended.', 'warning')
        return redirect(url_for('organizer.dashboard'))

    feedback_items = Feedback.query.filter_by(event_id=event_id).order_by(Feedback.submitted_at.desc()).all()
    avg_rating = db.session.query(func.avg(Feedback.rating)).filter_by(event_id=event_id).scalar() or 0

    return render_template('organizer/feedback_event.html',
                         event=event,
                         feedback_items=feedback_items,
                         avg_rating=round(avg_rating, 2))


@bp.route('/scan')
@organizer_required
def scan_select():
    """Event selection page for QR scanning - Mobile friendly"""
    organizer_id = session['user_id']
    
    # Get all approved events for this organizer
    events = Event.query.filter_by(
        organizer_id=organizer_id,
        status='approved'
    ).order_by(Event.date.desc()).all()
    
    # Add registration and attendance counts to each event
    for event in events:
        event.registrations_count = Registration.query.filter_by(event_id=event.event_id).count()
        event.attended_count = db.session.query(Attendance).join(Registration).filter(
            Registration.event_id == event.event_id
        ).count()
    
    return render_template('organizer/scan_select.html', events=events)


@bp.route('/scan-qr/<int:event_id>')
@organizer_required
def scan_qr(event_id):
    """QR code scanning interface - Production ready camera-based scanner"""
    organizer_id = session['user_id']
    
    event = Event.query.filter_by(
        event_id=event_id,
        organizer_id=organizer_id
    ).first_or_404()
    
    # Get registration and attendance counts
    registrations_count = Registration.query.filter_by(event_id=event_id).count()
    attended_count = db.session.query(Attendance).join(Registration).filter(
        Registration.event_id == event_id
    ).count()
    
    return render_template('organizer/scan_qr.html', 
                         event=event,
                         registrations_count=registrations_count,
                         attended_count=attended_count)


@bp.route('/event/<int:event_id>/stats')
@organizer_required
def get_event_stats(event_id):
    """Get real-time event statistics for AJAX updates"""
    organizer_id = session['user_id']
    
    event = Event.query.filter_by(
        event_id=event_id,
        organizer_id=organizer_id
    ).first()
    
    if not event:
        return jsonify({'success': False, 'message': 'Event not found'}), 404
    
    registrations_count = Registration.query.filter_by(event_id=event_id).count()
    attended_count = db.session.query(Attendance).join(Registration).filter(
        Registration.event_id == event_id
    ).count()
    
    return jsonify({
        'success': True,
        'registered': registrations_count,
        'attended': attended_count,
        'remaining': registrations_count - attended_count
    })


@bp.route('/scan/<int:event_id>')
@organizer_required
def scan_from_url(event_id):
    """Mark attendance from QR URL opened by external scanner apps"""
    organizer_id = session['user_id']
    event = Event.query.filter_by(
        event_id=event_id,
        organizer_id=organizer_id
    ).first_or_404()

    qr_code = (request.args.get('code') or '').strip()
    if not qr_code:
        flash('QR code is missing.', 'error')
        return redirect(url_for('organizer.view_event', event_id=event_id))

    qr_info = validate_qr_code(qr_code)
    if not qr_info:
        flash('Invalid QR code.', 'error')
        return redirect(url_for('organizer.view_event', event_id=event_id))

    if qr_info['event_id'] != event_id:
        flash('QR code is for a different event.', 'error')
        return redirect(url_for('organizer.view_event', event_id=event_id))

    registration = Registration.query.get(qr_info['registration_id'])
    if not registration or registration.event_id != event_id:
        flash('Registration not found for this event.', 'error')
        return redirect(url_for('organizer.view_event', event_id=event_id))

    try:
        event_start = datetime.combine(event.date, event.start_time)
        if datetime.now() < event_start:
            flash('Cannot mark attendance before event start time', 'error')
            return redirect(url_for('organizer.view_event', event_id=event_id))
        event_end = datetime.combine(event.date, event.end_time)
        if datetime.now() > (event_end + timedelta(days=3)):
            flash('Attendance marking window has expired (3 days).', 'error')
            return redirect(url_for('organizer.view_event', event_id=event_id))
    except Exception:
        pass

    existing = Attendance.query.filter_by(registration_id=registration.registration_id).first()
    if existing:
        flash('Attendance already marked for this student.', 'info')
        return redirect(url_for('organizer.view_event', event_id=event_id))

    attendance = Attendance(
        registration_id=registration.registration_id,
        scan_time=datetime.now(),
        scanned_by=organizer_id,
        status='present'
    )

    try:
        registration.attendance = attendance
        db.session.add(attendance)
        db.session.add(registration)
        db.session.commit()
    except Exception:
        db.session.rollback()
        flash('Failed to mark attendance due to database error.', 'error')
        return redirect(url_for('organizer.view_event', event_id=event_id))

    generate_certificate_for_student(registration.student_id, registration.event_id)
    flash(f'Attendance marked for {registration.student.full_name}', 'success')
    return redirect(url_for('organizer.view_event', event_id=event_id))


@bp.route('/validate-qr', methods=['POST'])
@organizer_required
def validate_qr():
    """Validate QR code and mark attendance (legacy endpoint)"""
    data = request.get_json()
    qr_code = data.get('qr_code')
    event_id = data.get('event_id')
    
    # Validate QR code format
    from utils.qr_utils import validate_qr_code
    qr_info = validate_qr_code(qr_code)
    
    if not qr_info:
        return jsonify({'success': False, 'message': 'Invalid QR code format'})
    
    # Find registration
    registration = Registration.query.filter_by(qr_code=qr_code).first()
    
    if not registration:
        return jsonify({'success': False, 'message': 'Registration not found'})
    
    # Verify event matches
    if registration.event_id != int(event_id):
        return jsonify({'success': False, 'message': 'QR code is for a different event'})

    # Disallow marking attendance before the event start time
    try:
        event_obj = Event.query.get(registration.event_id)
        event_start_dt = datetime.combine(event_obj.date, event_obj.start_time)
        if datetime.now() < event_start_dt:
            return jsonify({'success': False, 'message': 'Cannot mark attendance before event start time'})
        event_end_dt = datetime.combine(event_obj.date, event_obj.end_time)
        if datetime.now() > (event_end_dt + timedelta(days=3)):
            return jsonify({'success': False, 'message': 'Attendance window expired (3 days)'})
    except Exception:
        # If we cannot determine event time, allow normal processing to continue
        pass
    
    # Check if already scanned
    existing_attendance = Attendance.query.filter_by(
        registration_id=registration.registration_id
    ).first()
    
    if existing_attendance:
        return jsonify({
            'success': False, 
            'message': 'Already marked present',
            'student_name': registration.student.full_name,
            'scan_time': existing_attendance.scan_time.strftime('%Y-%m-%d %H:%M:%S')
        })
    
    # Mark attendance
    attendance = Attendance(
        registration_id=registration.registration_id,
        scan_time=datetime.now(),
        scanned_by=session['user_id'],
        status='present'
    )
    try:
        # Attach to registration to keep ORM relationship consistent
        registration.attendance = attendance
        db.session.add(attendance)
        db.session.add(registration)
        db.session.commit()
    except Exception:
        db.session.rollback()
        return jsonify({'success': False, 'message': 'Failed to mark attendance (DB error)'}), 500

    # Generate certificate after marking attendance
    generate_certificate_for_student(registration.student_id, registration.event_id)
    
    return jsonify({
        'success': True,
        'message': 'Attendance marked successfully',
        'student_name': registration.student.full_name,
        'student_email': registration.student.email
    })


# =============================================================================
# NEW API ENDPOINT: /api/scan-qr - Production-Ready QR Scanning API
# =============================================================================
@bp.route('/api/scan-qr', methods=['POST'])
@organizer_required
def api_scan_qr():
    """
    Production-ready API endpoint for QR code scanning.
    
    Input (JSON):
        - qr_code: The scanned QR code string (or URL containing ?code=...)
        - event_id: The event ID to validate against
    
    Output (JSON):
        - status: 'success' | 'duplicate' | 'invalid'
        - message: Human-readable message
        - student_name: Name of the student (if found)
        - student_email: Email of the student (if found)
        - event_name: Name of the event
        - timestamp: Time of scan
        - scan_time: Previous scan time (for duplicates)
    """
    # Validate request
    if not request.is_json:
        return jsonify({
            'status': 'invalid',
            'message': 'Request must be JSON'
        }), 400
    
    data = request.get_json()
    qr_code = (data.get('qr_code') or '').strip()
    event_id = data.get('event_id')
    
    # Validate required fields
    if not qr_code:
        return jsonify({
            'status': 'invalid',
            'message': 'QR code is required'
        }), 400
    
    if not event_id:
        return jsonify({
            'status': 'invalid',
            'message': 'Event ID is required'
        }), 400
    
    try:
        event_id = int(event_id)
    except (ValueError, TypeError):
        return jsonify({
            'status': 'invalid',
            'message': 'Invalid event ID'
        }), 400
    
    # Validate organizer owns this event
    organizer_id = session['user_id']
    event = Event.query.filter_by(event_id=event_id, organizer_id=organizer_id).first()
    
    if not event:
        return jsonify({
            'status': 'invalid',
            'message': 'Event not found or access denied'
        }), 403
    
    # Validate QR code format and extract data
    qr_info = validate_qr_code(qr_code)
    
    if not qr_info:
        return jsonify({
            'status': 'invalid',
            'message': 'Invalid QR code format. Please scan a valid registration QR code.',
            'event_name': event.title
        })
    
    # Find the registration by the original QR code string
    # First try exact match, then try by registration_id from parsed data
    registration = None
    
    # Extract the raw QR data if it was a URL
    raw_qr_code = qr_code
    if qr_code.startswith(('http://', 'https://')):
        from urllib.parse import urlparse, parse_qs
        parsed = urlparse(qr_code)
        params = parse_qs(parsed.query)
        if 'code' in params and params['code']:
            raw_qr_code = params['code'][0]
    
    # Try to find registration by QR code
    registration = Registration.query.filter_by(qr_code=raw_qr_code).first()
    
    # Fallback: try by registration_id from parsed QR info
    if not registration and qr_info.get('registration_id'):
        registration = Registration.query.get(qr_info['registration_id'])
    
    if not registration:
        return jsonify({
            'status': 'invalid',
            'message': 'Registration not found. This QR code may be invalid or expired.',
            'event_name': event.title
        })
    
    # Verify registration is for this event
    if registration.event_id != event_id:
        return jsonify({
            'status': 'invalid',
            'message': 'This QR code is for a different event.',
            'event_name': event.title
        })
    
    # Verify student exists
    student = registration.student
    if not student:
        return jsonify({
            'status': 'invalid',
            'message': 'Student record not found.',
            'event_name': event.title
        })
    
    # Check event timing constraints
    now = datetime.now()
    try:
        event_start = datetime.combine(event.date, event.start_time)
        event_end = datetime.combine(event.date, event.end_time)
        
        # Cannot scan before event starts
        if now < event_start:
            return jsonify({
                'status': 'invalid',
                'message': f'Event has not started yet. Scanning begins at {event.start_time.strftime("%H:%M")}.',
                'student_name': student.full_name,
                'event_name': event.title
            })
        
        # Allow scanning up to 3 days after event ends
        if now > (event_end + timedelta(days=3)):
            return jsonify({
                'status': 'invalid',
                'message': 'Attendance marking window has expired (3 days after event).',
                'student_name': student.full_name,
                'event_name': event.title
            })
    except Exception as e:
        # If time validation fails, continue with attendance marking
        current_app.logger.warning(f"Event time validation failed: {e}")
    
    # Check for duplicate attendance
    existing_attendance = Attendance.query.filter_by(
        registration_id=registration.registration_id
    ).first()
    
    if existing_attendance:
        return jsonify({
            'status': 'duplicate',
            'message': 'Attendance already recorded for this student.',
            'student_name': student.full_name,
            'student_email': student.email,
            'event_name': event.title,
            'scan_time': existing_attendance.scan_time.strftime('%Y-%m-%d %H:%M:%S')
        })
    
    # Mark attendance
    scan_time = datetime.now()
    attendance = Attendance(
        registration_id=registration.registration_id,
        scan_time=scan_time,
        scanned_by=organizer_id,
        status='present'
    )
    
    try:
        registration.attendance = attendance
        db.session.add(attendance)
        db.session.add(registration)
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Failed to mark attendance: {e}")
        return jsonify({
            'status': 'invalid',
            'message': 'Database error. Please try again.',
            'event_name': event.title
        }), 500
    
    # Generate certificate after successful attendance
    try:
        generate_certificate_for_student(student.user_id, event_id)
    except Exception as e:
        current_app.logger.warning(f"Certificate generation failed: {e}")
        # Don't fail the request - attendance was marked successfully
    
    return jsonify({
        'status': 'success',
        'message': 'Attendance marked successfully!',
        'student_name': student.full_name,
        'student_email': student.email,
        'event_name': event.title,
        'timestamp': scan_time.strftime('%Y-%m-%d %H:%M:%S')
    })


def generate_certificate_for_student(student_id, event_id):
    """Generate certificate after attendance is marked"""
    from models.models import Team
    
    # Check if certificate already exists
    existing_cert = Certificate.query.filter_by(
        student_id=student_id,
        event_id=event_id
    ).first()
    
    if existing_cert:
        return
    
    # Get student and event details
    student = User.query.get(student_id)
    event = Event.query.get(event_id)
    
    # Generate certificate filename
    filename = f"cert_{student_id}_{event_id}_{datetime.now().strftime('%Y%m%d%H%M%S')}.pdf"
    cert_path = os.path.join('static', 'uploads', 'certificates', filename)
    
    # Ensure directory exists
    os.makedirs(os.path.dirname(cert_path), exist_ok=True)
    
    # Check if student is part of a team for this event (for prize info)
    prize_text = None
    prize_template = None
    
    # Get the student's registration
    registration = Registration.query.filter_by(
        event_id=event_id,
        student_id=student_id
    ).first()
    
    if event.is_team_event:
        # For team events, check team prize
        if registration and registration.team:
            team = registration.team
            if team.prize_position and team.prize_position != 'Participant':
                # Build prize text
                if team.prize_title:
                    prize_text = f"{team.prize_position} - {team.prize_title}"
                else:
                    prize_text = f"{team.prize_position} Place"
                
                # Use team's prize-specific certificate template if assigned
                if team.prize_certificate_template_id:
                    prize_template = CertificateTemplate.query.get(team.prize_certificate_template_id)
    else:
        # For individual events, check individual prize on registration
        if registration and registration.prize_position and registration.prize_position != 'Participant':
            if registration.prize_title:
                prize_text = f"{registration.prize_position} - {registration.prize_title}"
            else:
                prize_text = f"{registration.prize_position} Place"
            
            # Use individual's prize-specific certificate template if assigned
            if registration.prize_certificate_template_id:
                prize_template = CertificateTemplate.query.get(registration.prize_certificate_template_id)
    
    # Determine which template to use
    # Priority: 1) Prize template (team or individual), 2) Event template, 3) Organizer default
    template = prize_template
    
    if not template and event.certificate_template_id:
        template = CertificateTemplate.query.filter_by(
            template_id=event.certificate_template_id,
            organizer_id=event.organizer_id
        ).first()

    if not template:
        template = CertificateTemplate.query.filter_by(
            organizer_id=event.organizer_id,
            is_default=True
        ).first()

    if template and template.image_url:
        template_path = os.path.join('static', template.image_url)
        positions = {}
        if template.positions:
            try:
                positions = json.loads(template.positions)
            except Exception:
                positions = {}
        if os.path.exists(template_path):
            generate_certificate_with_template(
                student_name=student.full_name,
                event_title=event.title,
                event_date=event.date.strftime('%B %d, %Y'),
                organizer_name=event.organizer.full_name,
                output_path=cert_path,
                template_path=template_path,
                positions=positions,
                prize_text=prize_text
            )
        else:
            generate_certificate(
                student_name=student.full_name,
                event_title=event.title,
                event_date=event.date.strftime('%B %d, %Y'),
                organizer_name=event.organizer.full_name,
                output_path=cert_path,
                prize_text=prize_text
            )
    else:
        generate_certificate(
            student_name=student.full_name,
            event_title=event.title,
            event_date=event.date.strftime('%B %d, %Y'),
            organizer_name=event.organizer.full_name,
            output_path=cert_path,
            prize_text=prize_text
        )
    
    # Save certificate record
    certificate = Certificate(
        student_id=student_id,
        event_id=event_id,
        certificate_url=f'uploads/certificates/{filename}'
    )
    db.session.add(certificate)
    db.session.commit()


@bp.route('/mark-attendance/<int:registration_id>', methods=['POST'])
@organizer_required
def mark_attendance(registration_id):
    """Manually mark attendance for a registration (used for online events)"""
    organizer_id = session['user_id']
    # Find registration and event
    registration = Registration.query.get_or_404(registration_id)
    event = Event.query.get_or_404(registration.event_id)

    # Ensure organizer owns the event
    if event.organizer_id != organizer_id:
        flash('Access denied', 'error')
        return redirect(url_for('organizer.view_event', event_id=event.event_id))

    # Only allow manual marking for online events (or still allow if organizer wants)
    # We'll allow it but prefer online check
    if (event.mode or '').lower() == 'offline':
        # still allow but warn
        flash('Manual marking is intended for online events. Proceeding.', 'warning')

    # Check if already marked
    existing = Attendance.query.join(Registration).filter(
        Attendance.registration_id == registration.registration_id
    ).first()
    if existing:
        flash('Attendance already marked for this registration', 'info')
        return redirect(url_for('organizer.view_event', event_id=event.event_id))

    # Prevent marking before event start time
    try:
        event_start = datetime.combine(event.date, event.start_time)
        if datetime.now() < event_start:
            flash('Cannot mark attendance before the event start time', 'error')
            return redirect(url_for('organizer.view_event', event_id=event.event_id))
        event_end = datetime.combine(event.date, event.end_time)
        if datetime.now() > (event_end + timedelta(days=3)):
            flash('Attendance marking window has expired (3 days).', 'error')
            return redirect(url_for('organizer.view_event', event_id=event.event_id))
    except Exception:
        # ignore and allow marking if we cannot compute start
        pass

    # Mark attendance
    attendance = Attendance(
        registration_id=registration.registration_id,
        scan_time=datetime.now(),
        scanned_by=organizer_id,
        status='present'
    )
    try:
        # Attach attendance to the registration relationship to keep ORM state consistent
        registration.attendance = attendance
        db.session.add(attendance)
        db.session.add(registration)
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        flash('Failed to mark attendance due to database error.', 'error')
        return redirect(url_for('organizer.view_event', event_id=event.event_id))

    # Refresh registration and attendance objects to ensure relationship is available
    try:
        db.session.refresh(registration)
        db.session.refresh(attendance)
    except Exception:
        # If refresh fails, ignore - the redirect will re-query on the next request
        pass

    # Generate certificate
    generate_certificate_for_student(registration.student_id, registration.event_id)

    flash('Attendance marked successfully', 'success')
    return redirect(url_for('organizer.view_event', event_id=event.event_id))
